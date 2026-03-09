"""Validate cleaned historical CSVs for structural integrity, DB consistency, and cross-file checks."""

import argparse
import csv
import json
import random
import sys
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

CLEANED_DIR = Path(__file__).parent.parent / "cleaned"
MAPPINGS_DIR = Path(__file__).parent.parent / "mappings"
HISTORICAL_DIR = Path(__file__).parent.parent / "historical"
OLDER_DIR = HISTORICAL_DIR / "older"

# Tier C/D offers that should have mystery CSVs
TIER_CD_OFFERS = [22, 23, 24, 26, 28, 30, 32, 34, 36, 38, 40, 42, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54]
TIER_D_CUTOFF = 48  # offers <= this are Tier D


class Severity(Enum):
    PASS = "PASS"
    WARN = "WARN"
    FAIL = "FAIL"
    SKIP = "SKIP"
    INFO = "INFO"


@dataclass
class CheckResult:
    name: str
    severity: Severity
    message: str
    detail: dict = field(default_factory=dict)


@dataclass
class OfferReport:
    offer_id: int
    tier: str
    checks: list[CheckResult] = field(default_factory=list)

    @property
    def fails(self) -> int:
        return sum(1 for c in self.checks if c.severity == Severity.FAIL)

    @property
    def warns(self) -> int:
        return sum(1 for c in self.checks if c.severity == Severity.WARN)

    @property
    def status(self) -> str:
        if self.fails:
            return "FAIL"
        if self.warns:
            return "WARN"
        return "PASS"


def _severity_color(sev: Severity) -> str:
    return {
        Severity.PASS: "\033[32m",
        Severity.WARN: "\033[33m",
        Severity.FAIL: "\033[31m",
        Severity.SKIP: "\033[90m",
        Severity.INFO: "\033[36m",
    }.get(sev, "")


RESET = "\033[0m"


def _offer_tier(offer_id: int) -> str:
    if offer_id >= 64:
        return "A" if offer_id >= 64 else "B"
    if offer_id >= 55:
        return "B"
    if offer_id >= 49:
        return "C"
    return "D"


def _parse_only_offers(raw: str) -> set[int]:
    ids = set()
    for part in raw.split(","):
        part = part.strip()
        if "-" in part:
            lo, hi = part.split("-", 1)
            ids.update(range(int(lo), int(hi) + 1))
        else:
            ids.add(int(part))
    return ids


def _load_csv(path: Path) -> tuple[list[str], list[dict]] | None:
    """Load CSV, return (headers, rows) or None on failure."""
    try:
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = list(reader)
        return headers, rows
    except Exception:
        return None


def _load_name_map(offer_id: int) -> dict | None:
    path = MAPPINGS_DIR / f"offer_{offer_id}_name_map.json"
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Tier 1: Structural integrity
# ---------------------------------------------------------------------------

def check_csv_loadable(path: Path) -> CheckResult:
    result = _load_csv(path)
    if result is None:
        return CheckResult("csv_loadable", Severity.FAIL, f"Cannot parse {path.name}")
    headers, rows = result
    if not headers or headers[0] != "id":
        return CheckResult("csv_loadable", Severity.FAIL,
                           f"First column is '{headers[0] if headers else '(empty)'}', expected 'id'")
    box_cols = [h for h in headers if h != "id"]
    return CheckResult("csv_loadable", Severity.PASS,
                       f"{len(rows)} items, {len(box_cols)} boxes",
                       {"items": len(rows), "boxes": len(box_cols)})


def check_unique_ids(rows: list[dict]) -> CheckResult:
    ids = []
    bad = []
    for row in rows:
        raw = row.get("id", "")
        try:
            v = int(raw)
            if v <= 0:
                bad.append(raw)
            ids.append(v)
        except (ValueError, TypeError):
            bad.append(raw)
    if bad:
        return CheckResult("unique_ids", Severity.FAIL,
                           f"Non-positive/non-integer IDs: {bad[:5]}")
    dupes = [x for x in set(ids) if ids.count(x) > 1]
    if dupes:
        return CheckResult("unique_ids", Severity.FAIL,
                           f"Duplicate IDs: {dupes[:10]}")
    return CheckResult("unique_ids", Severity.PASS,
                       f"All {len(ids)} IDs unique and positive")


def check_quantities_valid(rows: list[dict], box_cols: list[str]) -> CheckResult:
    negatives = []
    high = []
    for row in rows:
        item_id = row.get("id", "?")
        for col in box_cols:
            raw = row.get(col, "0")
            try:
                v = float(raw) if raw else 0
            except (ValueError, TypeError):
                negatives.append((item_id, col, raw))
                continue
            if v < 0:
                negatives.append((item_id, col, raw))
            elif v > 20:
                high.append((item_id, col, v))
    if negatives:
        return CheckResult("quantities_valid", Severity.FAIL,
                           f"{len(negatives)} negative/invalid quantities",
                           {"examples": negatives[:5]})
    if high:
        return CheckResult("quantities_valid", Severity.WARN,
                           f"{len(high)} quantities > 20 (max: {max(v for _, _, v in high)})",
                           {"examples": high[:5]})
    max_qty = 0
    for row in rows:
        for col in box_cols:
            try:
                v = float(row.get(col, "0") or "0")
                max_qty = max(max_qty, v)
            except (ValueError, TypeError):
                pass
    return CheckResult("quantities_valid", Severity.PASS,
                       f"All values non-negative (max: {max_qty:g})")


def check_no_duplicate_columns(headers: list[str]) -> CheckResult:
    seen = {}
    dupes = []
    for h in headers:
        if h in seen:
            dupes.append(h)
        seen[h] = True
    if dupes:
        return CheckResult("no_duplicate_columns", Severity.FAIL,
                           f"Duplicate columns: {dupes}")
    return CheckResult("no_duplicate_columns", Severity.PASS,
                       f"All {len(headers)} columns unique")


def check_all_zero_rows(rows: list[dict], box_cols: list[str]) -> CheckResult:
    zeros = []
    for row in rows:
        total = 0
        for col in box_cols:
            try:
                total += float(row.get(col, "0") or "0")
            except (ValueError, TypeError):
                pass
        if total == 0:
            zeros.append(row.get("id", "?"))
    if zeros:
        return CheckResult("all_zero_rows", Severity.WARN,
                           f"{len(zeros)} items with all-zero allocations: {zeros[:10]}")
    return CheckResult("all_zero_rows", Severity.PASS, "No all-zero rows")


def check_box_count(box_cols: list[str]) -> CheckResult:
    n = len(box_cols)
    if n < 2 or n > 20:
        return CheckResult("box_count", Severity.WARN,
                           f"Unusual box count: {n} (expected 2-20)")
    return CheckResult("box_count", Severity.PASS, f"{n} boxes")


def check_item_count(rows: list[dict]) -> CheckResult:
    n = len(rows)
    if n < 10 or n > 60:
        return CheckResult("item_count", Severity.WARN,
                           f"Unusual item count: {n} (expected 10-60)")
    return CheckResult("item_count", Severity.PASS, f"{n} items")


# ---------------------------------------------------------------------------
# Tier 2: Name-map validation
# ---------------------------------------------------------------------------

def check_ids_in_name_map(rows: list[dict], name_map: dict | None) -> CheckResult:
    if name_map is None:
        return CheckResult("ids_in_name_map", Severity.SKIP, "No name map file")
    map_ids = {entry["id"] for entry in name_map.values() if "id" in entry}
    csv_ids = set()
    for row in rows:
        try:
            csv_ids.add(int(row["id"]))
        except (ValueError, TypeError, KeyError):
            pass
    missing = csv_ids - map_ids
    if missing:
        return CheckResult("ids_in_name_map", Severity.WARN,
                           f"{len(missing)}/{len(csv_ids)} IDs not in name map: {sorted(missing)[:10]}")
    return CheckResult("ids_in_name_map", Severity.PASS,
                       f"{len(csv_ids)}/{len(csv_ids)} IDs covered")


def check_name_map_methods(name_map: dict | None) -> CheckResult:
    if name_map is None:
        return CheckResult("name_map_methods", Severity.SKIP, "No name map file")
    methods: dict[str, int] = {}
    for entry in name_map.values():
        m = entry.get("method", "unknown")
        methods[m] = methods.get(m, 0) + 1
    parts = [f"{count} {method}" for method, count in sorted(methods.items())]
    return CheckResult("name_map_methods", Severity.INFO,
                       f"Methods: {', '.join(parts)}")


# ---------------------------------------------------------------------------
# Tier 3: DB-backed validation
# ---------------------------------------------------------------------------

def check_ids_in_db(rows: list[dict], offer_id: int) -> CheckResult:
    try:
        from allocator.db import fetch_offer_parts_by_name
    except Exception as e:
        return CheckResult("ids_in_db", Severity.SKIP, f"DB import failed: {e}")

    try:
        include_deleted = offer_id <= TIER_D_CUTOFF
        parts = fetch_offer_parts_by_name(offer_id, include_deleted=include_deleted)
        if not parts and not include_deleted:
            parts = fetch_offer_parts_by_name(offer_id, include_deleted=True)
        db_ids = {v["id"] for v in parts.values()}
    except Exception as e:
        return CheckResult("ids_in_db", Severity.SKIP, f"DB query failed: {e}")

    csv_ids = set()
    for row in rows:
        try:
            csv_ids.add(int(row["id"]))
        except (ValueError, TypeError, KeyError):
            pass

    missing = csv_ids - db_ids
    suffix = " (include_deleted=True)" if offer_id <= TIER_D_CUTOFF else ""
    if missing:
        return CheckResult("ids_in_db", Severity.FAIL,
                           f"{len(missing)}/{len(csv_ids)} IDs not in DB{suffix}: {sorted(missing)[:10]}")
    return CheckResult("ids_in_db", Severity.PASS,
                       f"All {len(csv_ids)} IDs found{suffix}")


def check_box_names_parseable(box_cols: list[str]) -> CheckResult:
    try:
        from allocator.box_parser import classify_box
    except Exception as e:
        return CheckResult("box_names_parseable", Severity.SKIP, f"Import failed: {e}")

    sizes = {"small": 0, "medium": 0, "large": 0, "unknown": 0}
    unparsed = []
    for col in box_cols:
        _, size_tier, _ = classify_box(col)
        if size_tier is None:
            sizes["unknown"] += 1
            unparsed.append(col)
        else:
            sizes[size_tier] += 1

    parts = [f"{count} {size}" for size, count in sizes.items() if count > 0]
    if unparsed:
        return CheckResult("box_names_parseable", Severity.WARN,
                           f"{len(box_cols)} boxes ({', '.join(parts)}); unknown size: {unparsed[:5]}")
    return CheckResult("box_names_parseable", Severity.PASS,
                       f"{len(box_cols)} boxes: {', '.join(parts)}")


def check_value_per_box(rows: list[dict], box_cols: list[str], offer_id: int) -> CheckResult:
    try:
        from allocator.db import fetch_offer_parts_by_name
        from allocator.box_parser import classify_box
        from allocator.config import BOX_TIERS
    except Exception as e:
        return CheckResult("value_per_box", Severity.SKIP, f"Import failed: {e}")

    try:
        include_deleted = offer_id <= TIER_D_CUTOFF
        parts = fetch_offer_parts_by_name(offer_id, include_deleted=include_deleted)
        if not parts and not include_deleted:
            parts = fetch_offer_parts_by_name(offer_id, include_deleted=True)
        id_to_price = {v["id"]: v["price"] for v in parts.values()}
    except Exception as e:
        return CheckResult("value_per_box", Severity.SKIP, f"DB query failed: {e}")

    warnings = []
    pcts = []
    for col in box_cols:
        _, size_tier, _ = classify_box(col)
        if size_tier is None:
            continue
        target = BOX_TIERS.get(size_tier, {}).get("target_value")
        if not target:
            continue
        total = 0
        for row in rows:
            try:
                item_id = int(row["id"])
                qty = float(row.get(col, "0") or "0")
                price = id_to_price.get(item_id, 0)
                total += qty * price
            except (ValueError, TypeError):
                pass
        pct = total / target * 100 if target else 0
        pcts.append(pct)
        if pct < 50 or pct > 200:
            warnings.append(f"{col}: {pct:.0f}%")

    if not pcts:
        return CheckResult("value_per_box", Severity.SKIP, "No boxes with known size/target")
    lo, hi = min(pcts), max(pcts)
    if warnings:
        return CheckResult("value_per_box", Severity.WARN,
                           f"Range: {lo:.0f}-{hi:.0f}% of target; outliers: {warnings[:5]}")
    return CheckResult("value_per_box", Severity.PASS,
                       f"Range: {lo:.0f}-{hi:.0f}% of target")


# ---------------------------------------------------------------------------
# Tier 4: Cross-file consistency
# ---------------------------------------------------------------------------

def check_charity_no_overlap(mystery_cols: list[str], charity_cols: list[str]) -> CheckResult:
    overlap = set(mystery_cols) & set(charity_cols)
    if overlap:
        return CheckResult("charity_no_overlap", Severity.FAIL,
                           f"Shared columns: {sorted(overlap)}")
    return CheckResult("charity_no_overlap", Severity.PASS,
                       f"Mystery ({len(mystery_cols)} cols) and charity ({len(charity_cols)} cols) disjoint")


def check_charity_ids_reasonable(mystery_rows: list[dict], charity_rows: list[dict],
                                  name_map: dict | None) -> CheckResult:
    mystery_ids = {int(r["id"]) for r in mystery_rows if r.get("id")}
    charity_ids = {int(r["id"]) for r in charity_rows if r.get("id")}
    name_map_ids = {e["id"] for e in name_map.values()} if name_map else set()
    known_ids = mystery_ids | name_map_ids
    unknown = charity_ids - known_ids
    if unknown:
        return CheckResult("charity_ids_reasonable", Severity.WARN,
                           f"{len(unknown)} charity IDs not in mystery/name_map: {sorted(unknown)[:10]}")
    return CheckResult("charity_ids_reasonable", Severity.PASS,
                       f"All {len(charity_ids)} charity IDs are plausible")


def check_charity_qty_plausible(mystery_rows: list[dict], mystery_cols: list[str],
                                 charity_rows: list[dict], charity_cols: list[str]) -> CheckResult:
    mystery_totals: dict[str, float] = {}
    for row in mystery_rows:
        item_id = row.get("id", "?")
        total = 0
        for col in mystery_cols:
            try:
                total += float(row.get(col, "0") or "0")
            except (ValueError, TypeError):
                pass
        mystery_totals[item_id] = total

    outliers = []
    for row in charity_rows:
        item_id = row.get("id", "?")
        for col in charity_cols:
            try:
                qty = float(row.get(col, "0") or "0")
            except (ValueError, TypeError):
                continue
            m_total = mystery_totals.get(item_id, 0)
            if qty > 0 and m_total > 0 and qty > m_total * 3:
                outliers.append((item_id, col, qty, m_total))

    if outliers:
        examples = [f"ID {i}: charity={q:g} vs mystery_total={m:g}" for i, _, q, m in outliers[:5]]
        return CheckResult("charity_qty_plausible", Severity.WARN,
                           f"{len(outliers)} charity qtys >3x mystery total: {'; '.join(examples)}")
    return CheckResult("charity_qty_plausible", Severity.PASS, "Charity quantities plausible")


# ---------------------------------------------------------------------------
# Tier 5: LLM spot-check
# ---------------------------------------------------------------------------

def _extract_xlsx_rows(xlsx_path: Path, sample_ids: list[int]) -> tuple[list[list], list[list]] | None:
    """Read XLSX header + rows matching sample IDs. Returns (header_rows, data_rows) or None."""
    try:
        import openpyxl
        from allocator.clean_history import select_allocation_sheet
    except Exception:
        return None

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws, sheet_name, header_row = select_allocation_sheet(wb, 0)
        if ws is None:
            # Fallback: use first sheet, row 1
            ws = wb[wb.sheetnames[0]]
            header_row = 1

        # Read header
        headers = [cell.value for cell in list(ws.iter_rows(
            min_row=header_row, max_row=header_row))[0]]

        # Find ID column index
        id_col = None
        for i, h in enumerate(headers):
            if h is not None and str(h).strip() == "ID":
                id_col = i
                break

        # Read all data rows, keep those matching sample IDs
        sample_set = set(sample_ids)
        matched_rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if id_col is not None:
                try:
                    row_id = int(row[id_col])
                    if row_id in sample_set:
                        matched_rows.append(list(row))
                except (ValueError, TypeError, IndexError):
                    pass
            # If no ID column, collect first ~50 rows for context
            elif len(matched_rows) < 50:
                matched_rows.append(list(row))

        return [headers], matched_rows
    except Exception:
        return None


def check_llm_xlsx_verify(offer_id: int, rows: list[dict], box_cols: list[str], model: str = "haiku") -> CheckResult:
    try:
        from allocator.claude_cli import call_claude_cli
    except Exception as e:
        return CheckResult("llm_xlsx_verify", Severity.SKIP, f"Import failed: {e}")

    # Find XLSX
    xlsx_path = None
    for d in [OLDER_DIR, HISTORICAL_DIR]:
        for f in d.glob(f"offer_{offer_id}_*.*xlsx"):
            xlsx_path = f
            break
        if xlsx_path:
            break
    if not xlsx_path:
        return CheckResult("llm_xlsx_verify", Severity.SKIP, f"No XLSX found for offer {offer_id}")

    # Pick 5 random items
    sample = random.sample(rows, min(5, len(rows)))
    sample_ids = []
    csv_data = []
    for row in sample:
        item_id = row.get("id", "?")
        try:
            sample_ids.append(int(item_id))
        except (ValueError, TypeError):
            pass
        vals = {col: row.get(col, "0") for col in box_cols}
        csv_data.append({"id": item_id, **vals})

    # Extract matching rows from XLSX
    xlsx_data = _extract_xlsx_rows(xlsx_path, sample_ids)
    if xlsx_data is None:
        return CheckResult("llm_xlsx_verify", Severity.SKIP, f"Could not read XLSX for offer {offer_id}")
    header_rows, data_rows = xlsx_data
    if not data_rows:
        return CheckResult("llm_xlsx_verify", Severity.SKIP, "No matching rows found in XLSX")

    # Format XLSX data for prompt
    xlsx_text = f"XLSX headers: {header_rows[0]}\n"
    for r in data_rows:
        xlsx_text += f"  {r}\n"

    prompt = (
        f"Compare these two data sources for offer {offer_id}.\n\n"
        f"SOURCE 1 — Cleaned CSV (5 sample items with allocations per box):\n"
        f"{json.dumps(csv_data, indent=2)}\n\n"
        f"SOURCE 2 — Raw XLSX data (headers + matching rows from {xlsx_path.name}):\n"
        f"{xlsx_text}\n"
        f"The CSV box columns are: {box_cols}\n\n"
        f"Do the quantities in the CSV match the XLSX for these items? "
        f"Column names may differ between CSV and XLSX (the CSV uses cleaned names). "
        f"Focus on whether the numeric allocation values are correct.\n"
        f"If everything matches, say 'ALL MATCH'. Otherwise describe the discrepancies concisely."
    )

    result = call_claude_cli(prompt, model=model, timeout=120, lightweight=True)
    if result is None:
        return CheckResult("llm_xlsx_verify", Severity.SKIP, "Claude CLI call failed")

    if "ALL MATCH" in result.upper():
        return CheckResult("llm_xlsx_verify", Severity.PASS, "LLM verified 5 sample items match")
    return CheckResult("llm_xlsx_verify", Severity.WARN,
                       f"LLM found discrepancies: {result[:200]}")


# ---------------------------------------------------------------------------
# Main validation logic
# ---------------------------------------------------------------------------

def validate_offer(offer_id: int, *, use_db: bool = True, llm_check: bool = False, llm_model: str = "haiku") -> OfferReport:
    tier = _offer_tier(offer_id)
    report = OfferReport(offer_id=offer_id, tier=tier)

    mystery_path = CLEANED_DIR / f"offer_{offer_id}_mystery.csv"
    charity_path = CLEANED_DIR / f"offer_{offer_id}_charity.csv"

    # Tier 1: Structural
    load_result = check_csv_loadable(mystery_path)
    report.checks.append(load_result)
    if load_result.severity == Severity.FAIL:
        return report

    headers, rows = _load_csv(mystery_path)
    box_cols = [h for h in headers if h != "id"]

    report.checks.append(check_unique_ids(rows))
    report.checks.append(check_quantities_valid(rows, box_cols))
    report.checks.append(check_no_duplicate_columns(headers))
    report.checks.append(check_all_zero_rows(rows, box_cols))
    report.checks.append(check_box_count(box_cols))
    report.checks.append(check_item_count(rows))

    # Tier 2: Name-map
    name_map = _load_name_map(offer_id)
    report.checks.append(check_ids_in_name_map(rows, name_map))
    report.checks.append(check_name_map_methods(name_map))

    # Tier 3: DB-backed
    if use_db:
        report.checks.append(check_ids_in_db(rows, offer_id))
        report.checks.append(check_box_names_parseable(box_cols))
        report.checks.append(check_value_per_box(rows, box_cols, offer_id))
    else:
        for name in ("ids_in_db", "box_names_parseable", "value_per_box"):
            report.checks.append(CheckResult(name, Severity.SKIP, "Skipped (--no-db)"))

    # Tier 4: Cross-file (if charity CSV exists)
    if charity_path.exists():
        charity_result = _load_csv(charity_path)
        if charity_result:
            c_headers, c_rows = charity_result
            c_box_cols = [h for h in c_headers if h != "id"]
            report.checks.append(check_charity_no_overlap(box_cols, c_box_cols))
            report.checks.append(check_charity_ids_reasonable(rows, c_rows, name_map))
            report.checks.append(check_charity_qty_plausible(rows, box_cols, c_rows, c_box_cols))

    # Tier 5: LLM spot-check
    if llm_check:
        report.checks.append(check_llm_xlsx_verify(offer_id, rows, box_cols, model=llm_model))

    return report


def _discover_offer_ids(include_tier_a: bool = False) -> list[int]:
    """Find all offers that have mystery CSVs."""
    ids = set()
    for f in CLEANED_DIR.glob("offer_*_mystery.csv"):
        try:
            num = int(f.stem.split("_")[1])
            ids.add(num)
        except (ValueError, IndexError):
            pass
    if not include_tier_a:
        ids = {i for i in ids if i <= 54}
    return sorted(ids)


def print_report(report: OfferReport, verbose: bool = False):
    mystery_path = CLEANED_DIR / f"offer_{report.offer_id}_mystery.csv"
    charity_path = CLEANED_DIR / f"offer_{report.offer_id}_charity.csv"

    # Count items/boxes from csv_loadable detail
    csv_check = next((c for c in report.checks if c.name == "csv_loadable"), None)
    items = csv_check.detail.get("items", "?") if csv_check else "?"
    boxes = csv_check.detail.get("boxes", "?") if csv_check else "?"
    charity = "yes" if charity_path.exists() else "--"

    print(f"\n  OFFER {report.offer_id} (Tier {report.tier}) — {items} items, "
          f"{boxes} mystery boxes, charity: {charity}")

    for check in report.checks:
        color = _severity_color(check.severity)
        label = f"[{check.severity.value}]"
        if verbose or check.severity in (Severity.FAIL, Severity.WARN, Severity.INFO):
            print(f"  {color}{label:6s}{RESET} {check.name:24s} {check.message}")
        elif check.severity == Severity.PASS:
            print(f"  {color}{label:6s}{RESET} {check.name:24s} {check.message}")


def print_summary(reports: list[OfferReport]):
    print(f"\n  {'Offer':>5}  {'Tier':>4}  {'Items':>5}  {'Boxes':>5}  {'Charity':>7}  "
          f"{'Status':>6}  {'Fails':>5}  {'Warns':>5}")
    print(f"  {'─' * 5}  {'─' * 4}  {'─' * 5}  {'─' * 5}  {'─' * 7}  {'─' * 6}  {'─' * 5}  {'─' * 5}")

    for r in reports:
        csv_check = next((c for c in r.checks if c.name == "csv_loadable"), None)
        items = csv_check.detail.get("items", "?") if csv_check and csv_check.detail else "?"
        boxes = csv_check.detail.get("boxes", "?") if csv_check and csv_check.detail else "?"
        charity_path = CLEANED_DIR / f"offer_{r.offer_id}_charity.csv"
        charity = "yes" if charity_path.exists() else "--"

        status_color = {
            "PASS": "\033[32m", "WARN": "\033[33m", "FAIL": "\033[31m"
        }.get(r.status, "")
        print(f"  {r.offer_id:>5}  {r.tier:>4}  {items:>5}  {boxes:>5}  {charity:>7}  "
              f"{status_color}{r.status:>6}{RESET}  {r.fails:>5}  {r.warns:>5}")

    total = len(reports)
    passes = sum(1 for r in reports if r.status == "PASS")
    warns = sum(1 for r in reports if r.status == "WARN")
    fails = sum(1 for r in reports if r.status == "FAIL")
    print(f"\n  Overall: {total} offers | PASS: {passes} | WARN: {warns} | FAIL: {fails}")


def main():
    parser = argparse.ArgumentParser(description="Validate cleaned historical CSVs")
    parser.add_argument("--no-db", action="store_true", help="Skip DB-backed checks")
    parser.add_argument("--only-offers", type=str, default=None,
                        help="Comma-separated IDs/ranges (e.g. 22-48,50)")
    parser.add_argument("--include-tier-a", action="store_true",
                        help="Also validate Tier A/B offers")
    parser.add_argument("--llm-check", action="store_true",
                        help="Add LLM spot-check vs XLSX (run outside Claude Code)")
    parser.add_argument("--llm-model", type=str, default="haiku",
                        help="Model for --llm-check (default: haiku)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Show all check details")
    args = parser.parse_args()

    # Build offer list
    available = _discover_offer_ids(include_tier_a=args.include_tier_a)
    if args.only_offers:
        requested = _parse_only_offers(args.only_offers)
        offer_ids = sorted(requested & set(available))
    else:
        offer_ids = available

    if not offer_ids:
        print("No matching offers found with cleaned CSVs.")
        sys.exit(1)

    print(f"Validating {len(offer_ids)} offers: {offer_ids[0]}-{offer_ids[-1]}")

    reports = []
    for offer_id in offer_ids:
        report = validate_offer(offer_id, use_db=not args.no_db, llm_check=args.llm_check, llm_model=args.llm_model)
        print_report(report, verbose=args.verbose)
        reports.append(report)

    print_summary(reports)

    # Exit code: 1 if any FAILs
    if any(r.fails > 0 for r in reports):
        sys.exit(1)


if __name__ == "__main__":
    main()

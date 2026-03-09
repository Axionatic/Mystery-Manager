#!/usr/bin/env python3
"""
Compare LLM extraction outputs between methods side-by-side.

Reads production caches (mappings/offer_N_llm_extraction_{method}.json) and
matched CSVs (cleaned_llm/{method}/offer_N_mystery.csv) — no live LLM calls.

Usage:
    python3 scripts/compare_llm_outputs.py
    python3 scripts/compare_llm_outputs.py --methods haiku-whole,two-stage-smart-haiku
    python3 scripts/compare_llm_outputs.py --offers 22-30,45-54
    python3 scripts/compare_llm_outputs.py --detail 32
"""

import argparse
import json
import csv
import subprocess
import sys
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from rich.console import Console
from rich.table import Table
from rich import box as rich_box
from rich.progress import BarColumn, MofNCompleteColumn, Progress, TextColumn

ROOT = Path(__file__).parent.parent
MAPPINGS = ROOT / "mappings"
CLEANED_LLM = ROOT / "cleaned_llm"
LLM_COMPARISONS = ROOT / "llm_comparisons"

DEFAULT_METHODS = ["haiku-whole", "two-stage-smart-haiku"]

console = Console()


# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------

def _load_cache(offer_id: int, method: str) -> dict | None:
    path = MAPPINGS / f"offer_{offer_id}_llm_extraction_{method}.json"
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text())
        return data if data else None
    except (json.JSONDecodeError, OSError):
        return None


def _csv_matched(offer_id: int, method: str) -> int:
    path = CLEANED_LLM / method / f"offer_{offer_id}_mystery.csv"
    if not path.exists():
        return 0
    try:
        with path.open() as f:
            reader = csv.reader(f)
            rows = list(reader)
        # Subtract header row
        return max(0, len(rows) - 1)
    except OSError:
        return 0


# ---------------------------------------------------------------------------
# Tuple extraction and Jaccard similarity
# ---------------------------------------------------------------------------

def _to_tuples(cache_dict: dict) -> set[tuple]:
    """Extract (item_name_lower, box_index, qty) tuples from a raw cache dict."""
    tuples = set()
    for item in cache_dict.get("items", []):
        name = item.get("name", "").lower().strip()
        for i, qty in enumerate(item.get("allocations", [])):
            if isinstance(qty, (int, float)) and qty > 0:
                tuples.add((name, i, qty))
    return tuples


def _jaccard(set_a: set, set_b: set) -> float:
    if not set_a and not set_b:
        return 1.0
    return len(set_a & set_b) / len(set_a | set_b)


# ---------------------------------------------------------------------------
# Offer range parsing
# ---------------------------------------------------------------------------

def _parse_offer_ranges(s: str) -> set[int]:
    ids = set()
    for part in s.split(","):
        part = part.strip()
        if "-" in part:
            lo, hi = part.split("-", 1)
            ids.update(range(int(lo), int(hi) + 1))
        else:
            ids.add(int(part))
    return ids


# ---------------------------------------------------------------------------
# Discover covered offers
# ---------------------------------------------------------------------------

def _discover_offers(methods: list[str]) -> list[int]:
    """Union of offers that have a cache file for any of the given methods."""
    ids: set[int] = set()
    for p in MAPPINGS.glob("offer_*_llm_extraction_*.json"):
        parts = p.stem.split("_")
        # offer_N_llm_extraction_METHOD → parts[0]="offer", parts[1]=N
        try:
            oid = int(parts[1])
        except (IndexError, ValueError):
            continue
        # Check if this file's method is one we care about
        # filename: offer_N_llm_extraction_{method}
        # stem without offer_N_llm_extraction_ prefix
        prefix = f"offer_{oid}_llm_extraction_"
        method_in_file = p.stem[len(prefix):]
        if method_in_file in methods:
            ids.add(oid)
    return sorted(ids)


# ---------------------------------------------------------------------------
# Winner heuristic
# ---------------------------------------------------------------------------

def _winner(caches: list[dict | None], matched: list[int], items: list[int]) -> str:
    """
    Returns label like "A", "B", "=", or method letter for multiple methods.
    For two methods: A/B/=.
    """
    labels = [chr(65 + i) for i in range(len(caches))]  # A, B, C...

    failed = [c is None or not c.get("items") for c in caches]

    # One failed → other wins
    if all(failed):
        return "?"
    for i, f in enumerate(failed):
        if f:
            # Find first non-failed
            for j, f2 in enumerate(failed):
                if not f2:
                    return labels[j]

    # Both succeeded: more matched wins
    best_matched = max(matched)
    winners_m = [i for i, m in enumerate(matched) if m == best_matched]
    if len(winners_m) == 1:
        return labels[winners_m[0]]

    # Tie on matched: more raw items wins
    best_items = max(items)
    winners_i = [i for i, it in enumerate(items) if it == best_items]
    if len(winners_i) == 1:
        return labels[winners_i[0]]

    return "="


# ---------------------------------------------------------------------------
# Detail view
# ---------------------------------------------------------------------------

def _detail(offer_id: int, methods: list[str]) -> None:
    caches = [_load_cache(offer_id, m) for m in methods]

    console.print(f"\n[bold]Detail for offer {offer_id}[/bold]\n")

    if all(c is None for c in caches):
        console.print("[red]No cache found for any method.[/red]")
        return

    label_a, label_b = methods[0], methods[1] if len(methods) > 1 else "B"
    cache_a = caches[0] or {}
    cache_b = caches[1] if len(caches) > 1 else None
    cache_b = cache_b or {}

    tuples_a = _to_tuples(cache_a)
    tuples_b = _to_tuples(cache_b)

    names_a = {(name, box_i) for name, box_i, qty in tuples_a}
    names_b = {(name, box_i) for name, box_i, qty in tuples_b}

    only_a = tuples_a - tuples_b
    only_b = tuples_b - tuples_a

    # Group by item name (ignore box index for "only in" display)
    items_only_a = sorted({name for name, box_i, qty in only_a})
    items_only_b = sorted({name for name, box_i, qty in only_b})

    # Qty differs: same (name, box_i) but different qty
    name_box_a = {(name, box_i): qty for name, box_i, qty in tuples_a}
    name_box_b = {(name, box_i): qty for name, box_i, qty in tuples_b}
    common_keys = set(name_box_a) & set(name_box_b)
    qty_differs = [(k, name_box_a[k], name_box_b[k]) for k in sorted(common_keys)
                   if name_box_a[k] != name_box_b[k]]
    shared_count = len(tuples_a & tuples_b)

    console.print(f"[bold cyan]Only in {label_a}[/bold cyan] ({len(items_only_a)} unique item names):")
    for name in items_only_a:
        console.print(f"  • {name}")

    console.print(f"\n[bold cyan]Only in {label_b}[/bold cyan] ({len(items_only_b)} unique item names):")
    for name in items_only_b:
        console.print(f"  • {name}")

    console.print(f"\n[bold cyan]Qty differs[/bold cyan] ({len(qty_differs)}):")
    if qty_differs:
        t = Table(show_header=True, box=rich_box.SIMPLE)
        t.add_column("Item", style="dim")
        t.add_column("Box idx")
        t.add_column(label_a, justify="right")
        t.add_column(label_b, justify="right")
        for (name, box_i), qa, qb in qty_differs:
            t.add_row(name, str(box_i), str(qa), str(qb))
        console.print(t)
    else:
        console.print("  (none)")

    console.print(f"\n[bold cyan]Shared[/bold cyan]: {shared_count} tuples with identical (name, box_idx, qty)")


# ---------------------------------------------------------------------------
# Main summary
# ---------------------------------------------------------------------------

def _summary(methods: list[str], offer_ids: list[int]) -> dict[int, dict]:
    label_row = [chr(65 + i) for i in range(len(methods))]

    t = Table(title="LLM Extraction Comparison", box=rich_box.SIMPLE_HEAD, show_lines=False)
    t.add_column("Offer", justify="right", style="bold")
    for i, m in enumerate(methods):
        t.add_column(f"Sheet({label_row[i]})", max_width=16, no_wrap=True)
    t.add_column("Items " + "/".join(label_row), justify="right")
    t.add_column("Match " + "/".join(label_row), justify="right")
    if len(methods) == 2:
        t.add_column("Jaccard", justify="right")
    t.add_column("Winner", justify="center", no_wrap=True)

    # Aggregate stats
    agg_items = [[] for _ in methods]
    agg_matched = [[] for _ in methods]
    jaccards = []
    winner_counts: dict[str, int] = {}
    offer_data: dict[int, dict] = {}

    for oid in offer_ids:
        caches = [_load_cache(oid, m) for m in methods]
        matched = [_csv_matched(oid, m) for m in methods]
        items_counts = [len(c.get("items", [])) if c else 0 for c in caches]
        sheets = [
            (c.get("sheet_used", "") or "")[:18] if c else ""
            for c in caches
        ]

        tuples = [_to_tuples(c) if c else set() for c in caches]
        jac = _jaccard(tuples[0], tuples[1]) if len(methods) == 2 else None

        w = _winner(caches, matched, items_counts)

        # Flag low agreement
        flag = ""
        if jac is not None and jac < 0.60:
            flag = " \u2691"  # ⚑

        row = [str(oid)]
        for s in sheets:
            row.append(s)
        row.append("/".join(str(ic) if ic else "—" for ic in items_counts))
        row.append("/".join(str(m_count) if m_count else "—" for m_count in matched))
        if jac is not None:
            row.append(f"{jac*100:.0f}%")
        row.append(f"{w}{flag}")

        t.add_row(*row)

        # Aggregate
        if jac is not None:
            jaccards.append(jac)
        for i, (c, ic, m_count) in enumerate(zip(caches, items_counts, matched)):
            if c and ic:
                agg_items[i].append(ic)
            if m_count:
                agg_matched[i].append(m_count)
        winner_counts[w.rstrip()] = winner_counts.get(w.rstrip(), 0) + 1

        offer_data[oid] = {"caches": caches, "jac": jac, "matched": matched, "items": items_counts}

    console.print(t)

    # Aggregate summary
    console.print("\n[bold]Aggregate Summary[/bold]")
    agg_table = Table(box=rich_box.SIMPLE, show_header=True)
    agg_table.add_column("Method")
    agg_table.add_column("Offers", justify="right")
    agg_table.add_column("Avg items", justify="right")
    agg_table.add_column("Avg matched", justify="right")

    for i, m in enumerate(methods):
        n = len(agg_items[i])
        avg_items = f"{sum(agg_items[i])/n:.1f}" if n else "—"
        nm = len(agg_matched[i])
        avg_matched = f"{sum(agg_matched[i])/nm:.1f}" if nm else "—"
        agg_table.add_row(m, str(len(offer_ids)), avg_items, avg_matched)

    console.print(agg_table)

    if jaccards:
        avg_jac = sum(jaccards) / len(jaccards)
        console.print(f"Agreement (avg Jaccard): {avg_jac*100:.0f}%")

    # Winner breakdown
    parts = []
    for label in sorted(winner_counts):
        parts.append(f"{label} wins {winner_counts[label]}")
    console.print(f"Recommended per offer: {', '.join(parts)}")

    return offer_data


# ---------------------------------------------------------------------------
# Investigation helpers
# ---------------------------------------------------------------------------

def _sheet_to_text(path: Path, sheet_name: str, max_chars: int = 15000) -> str:
    """Read an XLSX sheet and return tab-separated text, truncated to max_chars."""
    try:
        import openpyxl
    except ImportError:
        return "(openpyxl not available)"
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        lines = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("\t".join(cells))
        wb.close()
    except Exception as e:
        return f"(error reading sheet: {e})"
    text = "\n".join(lines)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n... (truncated)"
    return text


def _find_xlsx(offer_id: int) -> Path | None:
    """Find the XLSX for the given offer across historical dirs."""
    from allocator.clean_history import discover_files
    historical = ROOT / "historical"
    older = ROOT / "historical" / "older"
    files = discover_files(historical, older)
    entry = files.get(offer_id)
    return entry[0] if entry else None


def _build_prompt(offer_id: int, methods: list[str], caches: list[dict | None]) -> str:
    xlsx_path = _find_xlsx(offer_id)

    # Collect sheet names from both caches
    sheets_seen: list[str] = []
    for c in caches:
        if c:
            sheet = c.get("sheet_used", "") or ""
            if sheet and sheet not in sheets_seen:
                sheets_seen.append(sheet)

    # Build sheet text block(s)
    sheet_sections = []
    if xlsx_path and sheets_seen:
        for sheet in sheets_seen:
            text = _sheet_to_text(xlsx_path, sheet)
            sheet_sections.append(f'**Sheet "{sheet}":**\n```\n{text}\n```')
    elif xlsx_path:
        # fallback: read active sheet
        text = _sheet_to_text(xlsx_path, "")
        sheet_sections.append(f"**Active sheet:**\n```\n{text}\n```")
    else:
        sheet_sections.append("*(XLSX not found)*")

    sheet_block = "\n\n".join(sheet_sections)

    # Build extraction blocks
    extraction_blocks = []
    labels = [chr(65 + i) for i in range(len(methods))]
    for label, method, cache in zip(labels, methods, caches):
        if cache:
            n_items = len(cache.get("items", []))
            # Count boxes from max allocations length
            n_boxes = max(
                (len(item.get("allocations", [])) for item in cache.get("items", [])),
                default=0,
            )
            cache_json = json.dumps(cache, indent=2)
        else:
            n_items = 0
            n_boxes = 0
            cache_json = "null"
        extraction_blocks.append(
            f"**Extraction {label} ({method}) — {n_items} items, {n_boxes} boxes:**\n"
            f"```json\n{cache_json}\n```"
        )

    extractions = "\n\n".join(extraction_blocks)
    label_a = f"{labels[0]} ({methods[0]})" if methods else "A"
    label_b = f"{labels[1]} ({methods[1]})" if len(methods) > 1 else "B"

    return f"""\
You are reviewing two automated LLM extractions for offer {offer_id} against the original spreadsheet.

**Original spreadsheet:**
{sheet_block}

{extractions}

Compare the two extractions against the original spreadsheet.
- What did each method get right or wrong (items, quantities, box assignments)?
- Note any specific mistakes made by {label_a} or {label_b}.
- Which extraction is more accurate overall, and why?

Be succinct."""


def _run_investigation(
    offer_id: int, methods: list[str], caches: list[dict | None], jac: float | None
) -> tuple[int, Path | None, str | None]:
    """Run claude -p sonnet investigation for one offer. Returns (offer_id, saved_path, error)."""
    LLM_COMPARISONS.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = LLM_COMPARISONS / f"offer_{offer_id}_{timestamp}.md"

    prompt = _build_prompt(offer_id, methods, caches)

    try:
        result = subprocess.run(
            ["claude", "-p", "--model", "sonnet"],
            input=prompt,
            capture_output=True,
            text=True,
            timeout=1800,
        )
        if result.returncode != 0:
            err = result.stderr.strip() or f"exit code {result.returncode}"
            return (offer_id, None, err)
        claude_output = result.stdout.strip()
    except subprocess.TimeoutExpired:
        return (offer_id, None, "timeout after 1800s")
    except FileNotFoundError:
        return (offer_id, None, "claude CLI not found")
    except Exception as e:
        return (offer_id, None, str(e))

    jac_str = f"{jac*100:.0f}%" if jac is not None else "n/a"
    ts_display = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    md = (
        f"# LLM Investigation: Offer {offer_id}\n"
        f"Date: {ts_display}\n"
        f"Methods: {methods[0] if methods else ''} vs {methods[1] if len(methods) > 1 else ''}\n"
        f"Jaccard: {jac_str}\n\n"
        f"## Claude's Analysis\n{claude_output}\n"
    )
    out_path.write_text(md)
    return (offer_id, out_path, None)


def _prompt_investigate(offer_ids: list[int], jaccards: dict[int, float | None]) -> list[int]:
    """Interactive TTY prompt. Returns selected offer IDs."""
    if not sys.stdin.isatty():
        return []

    console.print("\n[bold]Launch LLM investigations?[/bold]")
    console.print("  [N] No")
    console.print("  [A] All offers")
    console.print("  [I] Enter offer IDs (e.g. 22,30-35)")
    console.print("  [J] Jaccard <= X%  (will ask for threshold)")

    try:
        choice = input("Choice [N]: ").strip().upper() or "N"
    except (EOFError, KeyboardInterrupt):
        return []

    if choice == "N":
        return []
    if choice == "A":
        return list(offer_ids)
    if choice == "I":
        try:
            raw = input("Offer IDs: ").strip()
        except (EOFError, KeyboardInterrupt):
            return []
        selected = _parse_offer_ranges(raw) & set(offer_ids)
        return sorted(selected)
    if choice == "J":
        try:
            raw = input("Jaccard threshold % [50]: ").strip() or "50"
            threshold = float(raw) / 100.0
        except (EOFError, KeyboardInterrupt, ValueError):
            return []
        return sorted(
            oid for oid in offer_ids
            if jaccards.get(oid) is not None and jaccards[oid] <= threshold
        )
    return []


def _investigate_offers(targets: list[int], methods: list[str], offer_data: dict[int, dict]) -> None:
    """Run parallel investigations with Rich progress bar."""
    if not targets:
        return

    console.print(f"\n[bold]Investigating {len(targets)} offer(s)...[/bold]")
    saved_files: list[Path] = []

    with Progress(
        TextColumn("[bold]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        TextColumn("{task.fields[status]}"),
    ) as progress:
        task = progress.add_task("Investigating", total=len(targets), status="")

        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = {
                executor.submit(
                    _run_investigation,
                    oid,
                    methods,
                    offer_data[oid]["caches"],
                    offer_data[oid]["jac"],
                ): oid
                for oid in targets
            }

            for future in as_completed(futures):
                offer_id, saved_path, error = future.result()
                if error:
                    progress.update(task, advance=1, status=f"offer {offer_id}: ERROR — {error}")
                else:
                    saved_files.append(saved_path)
                    progress.update(
                        task, advance=1,
                        status=f"offer {offer_id}: saved {saved_path.name}",
                    )

    console.print("\n[bold]Saved files:[/bold]")
    for p in sorted(saved_files):
        console.print(f"  {p}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--methods", type=str, default=",".join(DEFAULT_METHODS),
                        help="Comma-separated method names (default: haiku-whole,two-stage-smart-haiku)")
    parser.add_argument("--offers", type=str, default=None,
                        help="Comma-separated offer IDs/ranges (e.g. '22-30,45-54')")
    parser.add_argument("--detail", type=int, default=None, metavar="OFFER_ID",
                        help="Show item-level diff for a single offer")
    args = parser.parse_args()

    methods = [m.strip() for m in args.methods.split(",") if m.strip()]
    if not methods:
        parser.error("--methods must name at least one method")

    if args.detail is not None:
        _detail(args.detail, methods)
        return

    # Build offer list
    all_covered = _discover_offers(methods)
    if args.offers:
        requested = _parse_offer_ranges(args.offers)
        offer_ids = sorted(requested & set(all_covered))
    else:
        offer_ids = all_covered

    if not offer_ids:
        console.print("[yellow]No offers found for the given methods/filters.[/yellow]")
        return

    offer_data = _summary(methods, offer_ids)

    if args.detail is None:
        jaccards = {oid: d["jac"] for oid, d in offer_data.items()}
        targets = _prompt_investigate(offer_ids, jaccards)
        _investigate_offers(targets, methods, offer_data)


if __name__ == "__main__":
    main()

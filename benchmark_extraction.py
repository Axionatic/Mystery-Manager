#!/usr/bin/env python3
"""
Benchmark LLM extraction strategies for non-standard historical workbooks.

Tests 6 strategies varying model (Sonnet vs Haiku), scope (whole workbook vs
per-tab), and staging (single-call vs two-stage) to find the fastest reliable
approach for extracting per-box allocation data.

Usage:
    # Top N largest Tier C/D workbooks, all strategies:
    python3 benchmark_extraction.py 5

    # Subset of strategies:
    python3 benchmark_extraction.py 3 --strategies sonnet-low,haiku-whole

    # Specific offers:
    python3 benchmark_extraction.py 0 --offers 32,40,52

    # Re-run ignoring cache:
    python3 benchmark_extraction.py 5 --force

Must be run outside Claude Code (nested claude -p restriction).
"""

import argparse
import json
import logging
import math
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from allocator.clean_history import (
    HISTORICAL_DIR,
    LLM_TARGET_OFFERS,
    OLDER_DIR,
    SKIP_OFFERS,
    discover_files,
)
from allocator.config import CHARITY_NAME
from allocator.sheet_analyzer import _TIER_A_EXAMPLE, _build_prompt

logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(name)s: %(message)s")
logger = logging.getLogger(__name__)

BENCHMARK_DIR = Path(__file__).parent / "benchmark_results"

ALL_STRATEGIES = [
    "sonnet-low",
    "haiku-whole",
    "haiku-per-tab",
    "sonnet-low-per-tab",
    "two-stage-haiku",
    "two-stage-sonnet",
    "two-stage-smart-haiku",
    "two-stage-smart-sonnet",
]


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class WorkbookInfo:
    """Pre-loaded workbook data ready for extraction."""

    offer_id: int
    path: Path
    size_kb: float
    sheet_names: list[str]
    sheets_text: dict[str, str]  # sheet_name -> stripped text
    workbook_text: str  # all sheets combined (as sheet_analyzer builds it)


@dataclass
class StrategyResult:
    """Result of running one strategy on one workbook."""

    raw_json: dict | None = None
    time_seconds: float = 0.0
    error: str | None = None
    notes: str = ""
    stage_times: dict[str, float] = field(default_factory=dict)
    raw_response: str | None = None

    @property
    def success(self) -> bool:
        return self.raw_json is not None and "boxes" in self.raw_json

    @property
    def n_boxes(self) -> int:
        if not self.success:
            return 0
        return len(self.raw_json["boxes"])

    @property
    def n_items(self) -> int:
        if not self.success:
            return 0
        return len(self.raw_json.get("items", []))

    @property
    def n_charity(self) -> int:
        if not self.success:
            return 0
        return len(self.raw_json.get("charity_boxes", []))


# ---------------------------------------------------------------------------
# Claude CLI wrapper (own copy — doesn't touch claude_cli.py)
# ---------------------------------------------------------------------------


def _call_claude(
    prompt: str,
    model: str = "sonnet",
    timeout: int = 300,
    effort: str | None = None,
    output_format: str = "text",
) -> tuple[str | None, float]:
    """
    Call claude -p with timing. Returns (output_or_None, elapsed_seconds).

    Always uses --tools "" --no-session-persistence for lightweight calls.
    If output_format=="json", appends --output-format json to the command.
    """
    cmd = ["claude", "-p", "--model", model, "--tools", "", "--no-session-persistence"]
    if effort:
        cmd.extend(["--effort", effort])
    if output_format == "json":
        cmd.extend(["--output-format", "json"])

    start = time.monotonic()
    try:
        result = subprocess.run(
            cmd,
            input=prompt,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        elapsed = time.monotonic() - start

        if result.returncode != 0:
            logger.warning(f"Claude CLI error (rc={result.returncode}): {result.stderr[:200]}")
            return None, elapsed

        return result.stdout.strip(), elapsed

    except subprocess.TimeoutExpired:
        elapsed = time.monotonic() - start
        logger.warning(f"Claude CLI timed out after {elapsed:.1f}s")
        return None, elapsed
    except FileNotFoundError:
        logger.error("claude binary not found — is it on PATH?")
        return None, 0.0
    except Exception as e:
        elapsed = time.monotonic() - start
        logger.warning(f"Claude CLI failed: {e}")
        return None, elapsed


# ---------------------------------------------------------------------------
# Workbook text formatting
# ---------------------------------------------------------------------------


def _format_sheet_stripped(
    ws, max_rows: int | None = None, formula_coords: set | None = None
) -> str:
    """
    Render worksheet as tab-separated text, skipping all-blank rows.

    Preserves original row numbers so LLM output can reference them.
    formula_coords: set of (row, col) 1-based tuples that contain Excel formulas.
    Those cells are rendered as "[F]" so the LLM knows they are computed values,
    not hand-entered allocation quantities.
    """
    if ws.max_row is None or ws.max_row < 1:
        return "(empty sheet)"

    n_rows = min(ws.max_row, max_rows) if max_rows else ws.max_row
    n_cols = ws.max_column or 1

    col_letters = [openpyxl.utils.get_column_letter(c) for c in range(1, n_cols + 1)]
    lines = ["Row\t" + "\t".join(col_letters)]

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=n_rows, values_only=True), 1
    ):
        # Skip all-blank rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        vals = []
        for c_idx, v in enumerate(row, 1):
            if formula_coords and (r_idx, c_idx) in formula_coords:
                vals.append("[F]")
            elif v is None:
                vals.append("")
            elif isinstance(v, float) and math.isfinite(v) and v == int(v):
                vals.append(str(int(v)))
            else:
                vals.append(str(v))
        lines.append(f"{r_idx}\t" + "\t".join(vals))

    return "\n".join(lines)


def _load_workbook_info(offer_id: int, path: Path) -> WorkbookInfo:
    """Load workbook and pre-compute all text representations.

    Loads twice: once with data_only=True for computed cell values, and once
    without to detect which cells contain Excel formulas. Formula cells are
    marked as "[F]" in the sheet text so the LLM knows they are computed values
    (e.g. price × quantity), not hand-entered mystery box allocation quantities.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)

    sheets_text = {}
    workbook_parts = []
    for sn in wb.sheetnames:
        ws = wb[sn]

        # Build formula coords from the formula workbook
        formula_coords: set[tuple[int, int]] = set()
        if sn in wb_formulas.sheetnames:
            ws_f = wb_formulas[sn]
            for row in ws_f.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        formula_coords.add((cell.row, cell.column))

        if ws.max_row is None or ws.max_row < 2:
            text = "(empty or 1-row sheet)"
        else:
            text = _format_sheet_stripped(ws, max_rows=100, formula_coords=formula_coords)
        sheets_text[sn] = text
        workbook_parts.append(
            f"### Sheet: {sn} ({ws.max_row or 0} rows x {ws.max_column or 0} cols)\n{text}\n"
        )

    return WorkbookInfo(
        offer_id=offer_id,
        path=path,
        size_kb=path.stat().st_size / 1024,
        sheet_names=wb.sheetnames,
        sheets_text=sheets_text,
        workbook_text="\n".join(workbook_parts),
    )


# ---------------------------------------------------------------------------
# Corner preview for smart two-stage classifier
# ---------------------------------------------------------------------------


def _extract_tab_corner(
    sheet_text: str,
    max_header_cols: int = 20,
    n_data_rows: int = 3,
    n_label_cols: int = 3,
) -> str:
    """
    Extract an L-shaped corner preview from a sheet's tab-separated text.

    Section A: first n_data_rows rows, full width capped at max_header_cols columns.
    Section B: all rows, first n_label_cols columns.

    The sheet_text is the output of _format_sheet_stripped, where the first line
    is a column-letter header (Row\\tA\\tB\\t...) and subsequent lines start with
    a row number.
    """
    lines = sheet_text.split("\n")
    if not lines:
        return "(empty)"

    header_line = lines[0]  # "Row\tA\tB\tC\t..."
    data_lines = lines[1:]

    def truncate_to_cols(line: str, n_cols: int) -> str:
        """Keep the row-number field plus the first n_cols column fields."""
        fields = line.split("\t")
        return "\t".join(fields[: n_cols + 1])

    # Section A: first n_data_rows rows, capped at max_header_cols columns
    section_a_lines = [truncate_to_cols(header_line, max_header_cols)]
    for line in data_lines:
        if not line.strip():
            continue
        fields = line.split("\t")
        try:
            row_num = int(fields[0])
        except (ValueError, IndexError):
            continue
        if row_num <= n_data_rows:
            section_a_lines.append(truncate_to_cols(line, max_header_cols))

    # Section B: all rows, first n_label_cols columns
    section_b_lines = [truncate_to_cols(header_line, n_label_cols)]
    for line in data_lines:
        if line.strip():
            section_b_lines.append(truncate_to_cols(line, n_label_cols))

    return (
        f"[First {n_data_rows} rows]\n"
        + "\n".join(section_a_lines)
        + f"\n\n[All rows, cols A–{'ABC'[n_label_cols - 1]}]\n"
        + "\n".join(section_b_lines)
    )


# ---------------------------------------------------------------------------
# Prompt builders
# ---------------------------------------------------------------------------


def _build_per_tab_prompt(offer_id: int, sheet_name: str, sheet_text: str) -> str:
    """Prompt for extracting from a single sheet tab."""
    return f"""\
You are analysing a single sheet from a fruit & veggie box business spreadsheet (offer {offer_id}).

## What allocation data looks like
Here is an example from a clean "Mystery Trello" sheet:

{_TIER_A_EXAMPLE}

Key features:
- "ID" column with DB item IDs, then one column per customer box (email or name), then {CHARITY_NAME} (charity)
- Data rows: item ID in first column, then integer quantities per box
- Box naming: emails (merged orders), "Md Name" / "Sm Name" (standalone), "{CHARITY_NAME}" (charity/donation)

## Your task
Does this sheet contain per-box mystery allocation data (how many of each item goes into each individual customer box)?

If YES: extract the data as JSON (format below).
If NO: reply with just the word `null`.

Important rules:
- Only extract individual customer box columns — NOT totals, subtotals, price columns, or supplier columns
- Each box column should represent ONE customer's box
- Items should be produce items (fruit/vegetables)
- Only include items with at least 1 non-zero allocation
- Empty/blank cells = 0
- Handle transposed layouts (items as columns, boxes as rows)
- Charity/donation boxes ({CHARITY_NAME}, donation emails) go in charity_boxes
- Cells marked [F] contain Excel formula results (e.g. price × quantity) — NOT hand-entered box allocations; skip any row/column where most allocation cells are [F] or non-integer decimals
- Valid allocation quantities are non-negative integers or 0.5 (occasionally used for large fruit or per-kg items like potatoes); any other decimal means the row/column is pricing data, not allocations

## Sheet: {sheet_name}

{sheet_text}

## Required output
Reply with ONLY valid JSON (no markdown fences) or the word `null`:

{{
  "sheet_used": "{sheet_name}",
  "data_orientation": "normal" or "transposed",
  "boxes": [
    {{"header": "original header", "name": "cleaned name", "size": "small|medium|large|unknown"}}
  ],
  "items": [
    {{"name": "item name", "allocations": [qty_per_box_1, qty_per_box_2, ...]}}
  ],
  "charity_boxes": [
    {{"header": "original header", "name": "cleaned name"}}
  ],
  "charity_items": [
    {{"name": "item name", "allocations": [qty_per_charity_box_1, ...]}}
  ],
  "notes": "any observations"
}}

The allocations arrays MUST match the length of their respective boxes array."""


def _build_tab_finder_prompt(offer_id: int, sheets_preview: dict[str, str]) -> str:
    """Prompt that sends first ~5 rows of each sheet to identify which has allocation data."""
    previews = []
    for sn, text in sheets_preview.items():
        # Take just the header + first 5 data rows
        lines = text.split("\n")
        preview = "\n".join(lines[:7])  # header row + up to 6 content rows
        previews.append(f"### Sheet: {sn}\n{preview}\n")

    all_previews = "\n".join(previews)

    return f"""\
You are looking at sheet previews (first few rows) from a fruit & veggie box spreadsheet (offer {offer_id}).

Which sheet contains the per-box mystery allocation data — i.e., a grid showing how many of each produce item goes into each individual customer box?

Look for sheets with:
- An "ID" or item name column
- Multiple columns for individual customer boxes (emails or names like "Md Name", "Sm Name")
- Integer quantity data in the grid

## Sheet previews

{all_previews}

Reply with ONLY the exact sheet name (nothing else). If none have allocation data, reply "NONE"."""


def _build_smart_classifier_prompt(offer_id: int, tab_corners: dict[str, str]) -> str:
    """
    Stage 1 prompt for two-stage-smart: sends L-shaped corner of each tab and
    asks Haiku to classify each one as allocation or not, with orientation info.
    """
    corners_text = []
    for sn, corner in tab_corners.items():
        corners_text.append(f"### Sheet: {sn}\n{corner}\n")

    all_corners = "\n".join(corners_text)

    return f"""\
You are analysing corner previews of sheets from a fruit & veggie mystery box allocation spreadsheet (offer {offer_id}).

Each sheet preview shows:
- [First 3 rows] — full width (up to 20 cols): reveals column header structure
- [All rows, cols A–C] — every row but only first 3 columns: reveals row-label patterns

## Known sheet orientations

**Normal orientation** (items as rows, boxes as columns):
- Col A has numeric item IDs (e.g. 123, 456)
- Row 1 has customer box headers: emails or names like "Md Name", "Sm Name", "Lge 1", "{CHARITY_NAME}"
- Cell values are small integers (quantities per box)

**Transposed orientation** (items as columns, boxes as rows):
- Col A has row labels including "Lge 1", "Med 1"–"Med 5", "small 1"–"small 5"
- Row 3 or nearby has item names as column headers
- Box-name rows appear after a metadata/header section

**Mixed** (shopping list + allocation columns on the right):
- Rows 1–3 have metadata labels like "Qty Sold", "Required Buy", "Overage", "JS Cost Ea", "RRP Ea"
- Allocation columns are far right (customer emails/names not visible in cols A–C)

**Irrelevant** (not allocation data):
- Customer refund tables (email, item name, quantity, price columns)
- Pure supplier price lists
- Summary/totals sheets without individual box breakdown

## Sheet corner previews

{all_corners}

## Required output
Reply with ONLY valid JSON (no markdown fences) — a list with one entry per sheet:

[
  {{
    "sheet": "exact sheet name",
    "orientation": "normal" or "transposed" or "mixed" or "unknown",
    "contains_allocation": true or false,
    "confidence": "high" or "medium" or "low",
    "data_start_row": <integer or null>,
    "notes": "brief description"
  }}
]

`data_start_row`: for transposed tabs, the row where box-name rows begin (after metadata).
For normal tabs, the first item data row (usually 2). Null if unknown."""


def _build_smart_extract_prompt(
    offer_id: int,
    sheet_name: str,
    sheet_text: str,
    classification: dict,
) -> str:
    """
    Stage 2 prompt for two-stage-smart: like _build_per_tab_prompt but injects
    the Stage 1 classification as explicit context to guide extraction.
    """
    orientation = classification.get("orientation", "unknown")
    data_start_row = classification.get("data_start_row")
    notes = classification.get("notes", "")
    confidence = classification.get("confidence", "unknown")

    if orientation == "normal":
        orient_instruction = "Items are in rows. Column A has IDs, row 1 has box headers (emails or box names). Read down the rows."
    elif orientation == "transposed":
        start_hint = f" around row {data_start_row}" if data_start_row else ""
        orient_instruction = (
            f"Items are columns. Box-name rows start{start_hint}. "
            "Rows before that are supplier/pricing metadata — skip them entirely. "
            "Each box-name row becomes one box; each item column has quantities."
        )
    elif orientation == "mixed":
        orient_instruction = (
            "This is a shopping-list sheet with mystery allocation columns to the far right. "
            "The first few columns are supplier data. Scan right for columns with email or box-name headers."
        )
    else:
        orient_instruction = "Orientation is unclear — check both row and column patterns."

    classification_block = f"""\
## Classification from pre-analysis
- Orientation: {orientation} (confidence: {confidence})
- Data starts at row: {data_start_row if data_start_row is not None else "unknown"}
- Notes: {notes}

{orient_instruction}"""

    return f"""\
You are analysing a single sheet from a fruit & veggie box business spreadsheet (offer {offer_id}).

{classification_block}

## What allocation data looks like
Here is an example from a clean "Mystery Trello" sheet:

{_TIER_A_EXAMPLE}

Key features:
- "ID" column with DB item IDs, then one column per customer box (email or name), then {CHARITY_NAME} (charity)
- Data rows: item ID in first column, then integer quantities per box
- Box naming: emails (merged orders), "Md Name" / "Sm Name" (standalone), "{CHARITY_NAME}" (charity/donation)

## Your task
Does this sheet contain per-box mystery allocation data (how many of each item goes into each individual customer box)?

If YES: extract the data as JSON (format below).
If NO: reply with just the word `null`.

Important rules:
- Only extract individual customer box columns — NOT totals, subtotals, price columns, or supplier columns
- Each box column should represent ONE customer's box
- Items should be produce items (fruit/vegetables)
- Only include items with at least 1 non-zero allocation
- Empty/blank cells = 0
- Handle transposed layouts (items as columns, boxes as rows)
- Charity/donation boxes ({CHARITY_NAME}, donation emails) go in charity_boxes
- Cells marked [F] contain Excel formula results (e.g. price × quantity) — NOT hand-entered box allocations; skip any row/column where most allocation cells are [F] or non-integer decimals
- Valid allocation quantities are non-negative integers or 0.5 (occasionally used for large fruit or per-kg items like potatoes); any other decimal means the row/column is pricing data, not allocations

## Sheet: {sheet_name}

{sheet_text}

## Required output
Reply with ONLY valid JSON (no markdown fences) or the word `null`:

{{
  "sheet_used": "{sheet_name}",
  "data_orientation": "normal" or "transposed",
  "boxes": [
    {{"header": "original header", "name": "cleaned name", "size": "small|medium|large|unknown"}}
  ],
  "items": [
    {{"name": "item name", "allocations": [qty_per_box_1, qty_per_box_2, ...]}}
  ],
  "charity_boxes": [
    {{"header": "original header", "name": "cleaned name"}}
  ],
  "charity_items": [
    {{"name": "item name", "allocations": [qty_per_charity_box_1, ...]}}
  ],
  "notes": "any observations"
}}

The allocations arrays MUST match the length of their respective boxes array."""


# ---------------------------------------------------------------------------
# JSON parsing helper
# ---------------------------------------------------------------------------


def _sanitize_allocations(raw_json: dict) -> dict:
    """
    Zero out allocation values that can't be valid mystery box quantities.

    Valid values: non-negative integers and 0.5 (used for per-kg items like
    potatoes, or large fruit like melons/pineapples). Any other decimal is
    almost certainly a formula-computed value (e.g. price × quantity).
    """

    def clean_val(v):
        if not isinstance(v, (int, float)):
            return 0
        if v == 0.5:
            return v
        if isinstance(v, float) and v != int(v):
            return 0  # non-0.5 decimal → formula result
        return v

    for key in ("items", "charity_items"):
        for item in raw_json.get(key, []):
            item["allocations"] = [clean_val(v) for v in item.get("allocations", [])]
    return raw_json


def _parse_json_response(raw: str | None) -> dict | None:
    """Parse JSON from LLM response, stripping markdown fences if present."""
    if raw is None:
        return None

    text = raw.strip()

    # Handle null responses (per-tab strategy)
    if text.lower() == "null":
        return None

    # Strip markdown fences
    if text.startswith("```"):
        lines = text.split("\n")
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines)

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as e:
        logger.debug(f"JSON parse failed ({e}): {text[:200]!r}")
        return None

    if isinstance(parsed, dict):
        _sanitize_allocations(parsed)
    return parsed


def _parse_claude_json_envelope(raw: str) -> str | None:
    """
    Parse the --output-format json envelope from claude CLI.

    The envelope looks like:
      {"type": "result", "subtype": "success", "is_error": false, "result": "<llm text>", ...}

    Returns the inner result string, or None on failure.
    """
    try:
        envelope = json.loads(raw)
        if envelope.get("is_error"):
            logger.warning(f"Claude returned error: {envelope.get('result', '')[:200]}")
            return None
        return envelope.get("result")
    except (json.JSONDecodeError, AttributeError):
        logger.debug(f"Envelope parse failed: {raw[:200]!r}")
        return None


# ---------------------------------------------------------------------------
# Sheet name matching for two-stage strategies
# ---------------------------------------------------------------------------


def _match_sheet_name(returned: str, available: list[str]) -> str | None:
    """
    Fuzzy-match a returned sheet name against available sheet names.

    Tries: exact → case-insensitive → substring match.
    """
    returned = returned.strip().strip("'\"")

    # Exact
    if returned in available:
        return returned

    # Case-insensitive
    lower_map = {s.lower(): s for s in available}
    if returned.lower() in lower_map:
        return lower_map[returned.lower()]

    # Substring (returned is substring of available, or vice versa)
    ret_lower = returned.lower()
    for s in available:
        if ret_lower in s.lower() or s.lower() in ret_lower:
            return s

    return None


# ---------------------------------------------------------------------------
# Strategy implementations
# ---------------------------------------------------------------------------


def run_sonnet_low(offer_id: int, info: WorkbookInfo) -> StrategyResult:
    """Whole workbook → Sonnet with --effort low."""
    prompt = _build_prompt(offer_id, info.workbook_text)
    raw, elapsed = _call_claude(prompt, model="sonnet", timeout=600, effort="low")
    parsed = _parse_json_response(raw)
    return StrategyResult(
        raw_json=parsed,
        time_seconds=elapsed,
        error=None if parsed else "parse_failed" if raw else "call_failed",
    )


def run_haiku_whole(offer_id: int, info: WorkbookInfo) -> StrategyResult:
    """Whole workbook → Haiku."""
    prompt = _build_prompt(offer_id, info.workbook_text)
    raw, elapsed = _call_claude(prompt, model="haiku", timeout=600)
    parsed = _parse_json_response(raw)
    return StrategyResult(
        raw_json=parsed,
        time_seconds=elapsed,
        error=None if parsed else "parse_failed" if raw else "call_failed",
    )


def _run_per_tab(
    offer_id: int, info: WorkbookInfo, model: str, effort: str | None
) -> StrategyResult:
    """Run extraction on each tab in parallel, pick the one with the most items."""
    if not info.sheet_names:
        return StrategyResult(time_seconds=0, error="no_sheets")

    futures = {}
    tab_times = {}
    total_start = time.monotonic()

    with ThreadPoolExecutor(max_workers=max(1, len(info.sheet_names))) as pool:
        for sn in info.sheet_names:
            prompt = _build_per_tab_prompt(offer_id, sn, info.sheets_text[sn])
            fut = pool.submit(_call_claude, prompt, model, 240, effort)
            futures[fut] = sn

        results_by_sheet = {}
        for fut in as_completed(futures):
            sn = futures[fut]
            raw, elapsed = fut.result()
            tab_times[sn] = elapsed
            parsed = _parse_json_response(raw)
            if parsed and "items" in parsed:
                results_by_sheet[sn] = parsed

    total_elapsed = time.monotonic() - total_start

    if not results_by_sheet:
        return StrategyResult(
            time_seconds=total_elapsed,
            error="no_tabs_matched",
            notes=f"{len(info.sheet_names)} tabs, 0 hits",
            stage_times=tab_times,
        )

    # Pick tab with most items
    best_sheet = max(results_by_sheet, key=lambda s: len(results_by_sheet[s].get("items", [])))
    best = results_by_sheet[best_sheet]
    hits = len(results_by_sheet)

    return StrategyResult(
        raw_json=best,
        time_seconds=total_elapsed,
        notes=f"{len(info.sheet_names)} tabs, {hits} hit{'s' if hits != 1 else ''}",
        stage_times=tab_times,
    )


def run_haiku_per_tab(offer_id: int, info: WorkbookInfo) -> StrategyResult:
    """Each tab → Haiku in parallel, pick best."""
    return _run_per_tab(offer_id, info, model="haiku", effort=None)


def run_sonnet_low_per_tab(offer_id: int, info: WorkbookInfo) -> StrategyResult:
    """Each tab → Sonnet --effort low in parallel, pick best."""
    return _run_per_tab(offer_id, info, model="sonnet", effort="low")


def _run_two_stage(
    offer_id: int, info: WorkbookInfo, extract_model: str, extract_effort: str | None
) -> StrategyResult:
    """Stage 1: Haiku finds the tab. Stage 2: extract from that tab only."""
    stage_times = {}

    # Stage 1: tab finder via Haiku
    finder_prompt = _build_tab_finder_prompt(offer_id, info.sheets_text)
    raw_name, s1_time = _call_claude(finder_prompt, model="haiku", timeout=120)
    stage_times["stage1_find"] = s1_time

    if raw_name is None:
        return StrategyResult(
            time_seconds=s1_time,
            error="stage1_failed",
            stage_times=stage_times,
        )

    # Check for NONE sentinel (LLM says no allocation sheet exists)
    if raw_name.strip().upper() == "NONE":
        return StrategyResult(
            time_seconds=s1_time,
            error="no_allocation_tab",
            stage_times=stage_times,
        )

    # Match returned name
    matched = _match_sheet_name(raw_name, info.sheet_names)
    if matched is None:
        return StrategyResult(
            time_seconds=s1_time,
            error=f"sheet_not_found: {raw_name!r}",
            notes=f"available: {info.sheet_names}",
            stage_times=stage_times,
        )

    # Stage 2: extract from matched sheet
    extract_prompt = _build_per_tab_prompt(offer_id, matched, info.sheets_text[matched])
    raw, s2_time = _call_claude(extract_prompt, extract_model, 360, extract_effort)
    stage_times["stage2_extract"] = s2_time

    total = s1_time + s2_time
    parsed = _parse_json_response(raw)

    return StrategyResult(
        raw_json=parsed,
        time_seconds=total,
        error=None if parsed else "stage2_parse_failed" if raw else "stage2_call_failed",
        notes=f"tab={matched}, stage1={s1_time:.1f}s",
        stage_times=stage_times,
        raw_response=raw if parsed is None else None,
    )


def run_two_stage_haiku(offer_id: int, info: WorkbookInfo) -> StrategyResult:
    """Haiku finds tab → Haiku extracts."""
    return _run_two_stage(offer_id, info, extract_model="haiku", extract_effort=None)


def run_two_stage_sonnet(offer_id: int, info: WorkbookInfo) -> StrategyResult:
    """Haiku finds tab → Sonnet --effort low extracts."""
    return _run_two_stage(offer_id, info, extract_model="sonnet", extract_effort="low")


_CONFIDENCE_RANK = {"high": 0, "medium": 1, "low": 2}
_ORIENTATION_RANK = {"normal": 0, "transposed": 1, "mixed": 2, "unknown": 3}


def _run_two_stage_smart(
    offer_id: int,
    info: WorkbookInfo,
    extract_model: str,
    extract_effort: str | None,
) -> StrategyResult:
    """
    Stage 1: Haiku classifies each tab via corner preview (--output-format json).
    Stage 2: Extract from the best tab using classification context.
    Falls back to haiku-whole if Stage 1 finds no allocation tab.
    """
    stage_times: dict[str, float] = {}

    # Stage 1: build corner previews and classify
    tab_corners = {sn: _extract_tab_corner(info.sheets_text[sn]) for sn in info.sheet_names}
    classifier_prompt = _build_smart_classifier_prompt(offer_id, tab_corners)
    raw_envelope, s1_time = _call_claude(
        classifier_prompt, model="haiku", timeout=120, output_format="json"
    )
    stage_times["stage1_classify"] = s1_time

    if raw_envelope is None:
        return StrategyResult(
            time_seconds=s1_time,
            error="stage1_call_failed",
            stage_times=stage_times,
        )

    inner_text = _parse_claude_json_envelope(raw_envelope)
    if inner_text is None:
        return StrategyResult(
            time_seconds=s1_time,
            error="stage1_envelope_parse_failed",
            stage_times=stage_times,
        )

    # Strip markdown fences if Haiku wrapped the JSON
    inner_stripped = inner_text.strip()
    if inner_stripped.startswith("```"):
        lines = inner_stripped.split("\n")
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        inner_stripped = "\n".join(lines)

    try:
        classifications = json.loads(inner_stripped)
        if not isinstance(classifications, list):
            raise ValueError("expected list")
    except (json.JSONDecodeError, ValueError) as e:
        logger.debug(f"Stage 1 JSON parse failed ({e}): {inner_stripped[:300]!r}")
        return StrategyResult(
            time_seconds=s1_time,
            error="stage1_json_parse_failed",
            stage_times=stage_times,
        )

    # Select best allocation tab: highest confidence, prefer normal > transposed > mixed
    candidates = [c for c in classifications if c.get("contains_allocation")]
    if not candidates:
        # Fallback to haiku-whole
        fallback = run_haiku_whole(offer_id, info)
        fallback.stage_times["stage1_classify"] = s1_time
        fallback.notes = f"fallback=haiku-whole (no allocation tab found by stage1)"
        return fallback

    best = min(
        candidates,
        key=lambda c: (
            _CONFIDENCE_RANK.get(c.get("confidence", "low"), 3),
            _ORIENTATION_RANK.get(c.get("orientation", "unknown"), 4),
        ),
    )

    # Fuzzy-match the sheet name
    matched = _match_sheet_name(best.get("sheet", ""), info.sheet_names)
    if matched is None:
        fallback = run_haiku_whole(offer_id, info)
        fallback.stage_times["stage1_classify"] = s1_time
        fallback.notes = f"fallback=haiku-whole (sheet {best.get('sheet')!r} not matched)"
        return fallback

    # Stage 2: extract using classification context
    extract_prompt = _build_smart_extract_prompt(
        offer_id, matched, info.sheets_text[matched], best
    )
    raw, s2_time = _call_claude(
        extract_prompt, model=extract_model, timeout=360, effort=extract_effort
    )
    stage_times["stage2_extract"] = s2_time

    total = s1_time + s2_time
    parsed = _parse_json_response(raw)
    orientation = best.get("orientation", "unknown")
    confidence = best.get("confidence", "unknown")

    return StrategyResult(
        raw_json=parsed,
        time_seconds=total,
        error=None if parsed else "stage2_parse_failed" if raw else "stage2_call_failed",
        notes=f"tab={matched}, orient={orientation}, conf={confidence}, stage1={s1_time:.1f}s",
        stage_times=stage_times,
        raw_response=raw if parsed is None else None,
    )


def run_two_stage_smart_haiku(offer_id: int, info: WorkbookInfo) -> StrategyResult:
    """Haiku classifies tabs (corner preview) → Haiku extracts with context."""
    return _run_two_stage_smart(offer_id, info, extract_model="haiku", extract_effort=None)


def run_two_stage_smart_sonnet(offer_id: int, info: WorkbookInfo) -> StrategyResult:
    """Haiku classifies tabs (corner preview) → Sonnet --effort low extracts with context."""
    return _run_two_stage_smart(offer_id, info, extract_model="sonnet", extract_effort="low")


STRATEGY_RUNNERS = {
    "sonnet-low": run_sonnet_low,
    "haiku-whole": run_haiku_whole,
    "haiku-per-tab": run_haiku_per_tab,
    "sonnet-low-per-tab": run_sonnet_low_per_tab,
    "two-stage-haiku": run_two_stage_haiku,
    "two-stage-sonnet": run_two_stage_sonnet,
    "two-stage-smart-haiku": run_two_stage_smart_haiku,
    "two-stage-smart-sonnet": run_two_stage_smart_sonnet,
}


# ---------------------------------------------------------------------------
# Caching
# ---------------------------------------------------------------------------


def _cache_path(offer_id: int, strategy: str) -> Path:
    return BENCHMARK_DIR / f"offer_{offer_id}_{strategy}.json"


def _load_cached(offer_id: int, strategy: str) -> StrategyResult | None:
    path = _cache_path(offer_id, strategy)
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text())
        if not isinstance(data, dict):
            return None
        return StrategyResult(
            raw_json=data.get("raw_json"),
            time_seconds=data.get("time_seconds", 0),
            error=data.get("error"),
            notes=data.get("notes", ""),
            stage_times=data.get("stage_times", {}),
            raw_response=data.get("raw_response"),
        )
    except (json.JSONDecodeError, AttributeError, TypeError):
        return None


def _save_cached(offer_id: int, strategy: str, result: StrategyResult):
    try:
        BENCHMARK_DIR.mkdir(parents=True, exist_ok=True)
        raw_resp = result.raw_response
        data = {
            "offer_id": offer_id,
            "strategy": strategy,
            "raw_json": result.raw_json,
            "time_seconds": result.time_seconds,
            "error": result.error,
            "notes": result.notes,
            "stage_times": result.stage_times,
            "raw_response": raw_resp[:2000] if raw_resp else None,
        }
        _cache_path(offer_id, strategy).write_text(json.dumps(data, indent=2))
    except Exception as e:
        logger.warning(f"Cache write failed for offer {offer_id}/{strategy}: {e}")


# ---------------------------------------------------------------------------
# Agreement scoring
# ---------------------------------------------------------------------------


def _extract_tuples(result: StrategyResult) -> set[tuple]:
    """Extract (item_name_lower, box_index, qty) tuples from a result."""
    if not result.success:
        return set()
    tuples = set()
    for item in result.raw_json.get("items", []):
        name = item.get("name", "").lower().strip()
        for i, qty in enumerate(item.get("allocations", [])):
            if isinstance(qty, (int, float)) and qty > 0:
                tuples.add((name, i, qty))
    return tuples


def _jaccard(set_a: set, set_b: set) -> float:
    """Jaccard similarity between two sets."""
    if not set_a and not set_b:
        return 1.0
    return len(set_a & set_b) / len(set_a | set_b)


# ---------------------------------------------------------------------------
# Workbook selection
# ---------------------------------------------------------------------------


def select_workbooks(n: int, offer_ids: set[int] | None = None) -> list[WorkbookInfo]:
    """
    Find workbooks for LLM target offers, sorted by file size descending.

    Args:
        n: Number of largest workbooks to return (0 = all)
        offer_ids: Override offer set (default: LLM_TARGET_OFFERS minus SKIP_OFFERS)
    """
    target_offers = offer_ids if offer_ids else (LLM_TARGET_OFFERS - SKIP_OFFERS)
    files = discover_files(HISTORICAL_DIR, OLDER_DIR)

    workbooks = []
    for oid in sorted(target_offers):
        if oid not in files:
            print(f"  Warning: no file found for offer {oid}")
            continue
        path, _source = files[oid]
        workbooks.append((oid, path, path.stat().st_size))

    # Sort by size descending
    workbooks.sort(key=lambda x: x[2], reverse=True)

    if n > 0:
        workbooks = workbooks[:n]

    print(f"Loading {len(workbooks)} workbook(s)...")
    infos = []
    for oid, path, _size in workbooks:
        try:
            info = _load_workbook_info(oid, path)
        except Exception as e:
            print(f"  Offer {oid}: SKIP ({e})")
            continue
        print(f"  Offer {oid}: {info.size_kb:.0f} KB, {len(info.sheet_names)} sheets")
        infos.append(info)

    return infos


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------


def print_per_workbook_table(
    offer_id: int, results: dict[str, StrategyResult], strategies: list[str]
):
    """Print results table for a single workbook."""
    from rich import box as rich_box
    from rich.console import Console
    from rich.table import Table

    console = Console()
    table = Table(
        title=f"Offer {offer_id}",
        box=rich_box.SIMPLE_HEAD,
        title_style="bold",
    )
    table.add_column("Strategy", min_width=20)
    table.add_column("Time", justify="right", width=8)
    table.add_column("OK?", justify="center", width=4)
    table.add_column("Boxes", justify="right", width=6)
    table.add_column("Items", justify="right", width=6)
    table.add_column("Charity", justify="right", width=8)
    table.add_column("Notes", min_width=20)

    for strat in strategies:
        r = results.get(strat)
        if r is None:
            continue
        ok = "[green]Y[/green]" if r.success else "[red]N[/red]"
        time_str = f"{r.time_seconds:.1f}s"
        notes = r.notes
        if r.error and not r.success:
            notes = f"[red]{r.error}[/red]"
        elif notes and "cached" in notes:
            notes = f"[dim]{notes}[/dim]"
        table.add_row(
            strat,
            time_str,
            ok,
            str(r.n_boxes) if r.success else "-",
            str(r.n_items) if r.success else "-",
            str(r.n_charity) if r.success else "-",
            notes,
        )

    console.print()
    console.print(table)


def print_aggregate_table(
    all_results: dict[int, dict[str, StrategyResult]], strategies: list[str]
):
    """Print aggregate results across all workbooks with agreement scoring."""
    from rich import box as rich_box
    from rich.console import Console
    from rich.table import Table

    console = Console()
    n_workbooks = len(all_results)

    table = Table(
        title=f"Aggregate Results  ({n_workbooks} workbooks)",
        box=rich_box.SIMPLE_HEAD,
        title_style="bold",
    )
    table.add_column("Strategy", min_width=20)
    table.add_column("Avg Time", justify="right", width=10)
    table.add_column("Success", justify="right", width=8)
    table.add_column("Avg Boxes", justify="right", width=10)
    table.add_column("Avg Items", justify="right", width=10)
    table.add_column("Agreement", justify="right", width=10)

    # Use sonnet-low as reference for agreement
    ref_strategy = "sonnet-low"
    ref_tuples = {}
    for oid, results in all_results.items():
        if ref_strategy in results and results[ref_strategy].success:
            ref_tuples[oid] = _extract_tuples(results[ref_strategy])

    for strat in strategies:
        times = []
        successes = 0
        box_counts = []
        item_counts = []
        agreements = []

        for oid, results in all_results.items():
            r = results.get(strat)
            if r is None:
                continue
            times.append(r.time_seconds)
            if r.success:
                successes += 1
                box_counts.append(r.n_boxes)
                item_counts.append(r.n_items)
                if oid in ref_tuples and strat != ref_strategy:
                    agreements.append(_jaccard(_extract_tuples(r), ref_tuples[oid]))

        avg_time = sum(times) / len(times) if times else 0
        avg_boxes = sum(box_counts) / len(box_counts) if box_counts else 0
        avg_items = sum(item_counts) / len(item_counts) if item_counts else 0

        if strat == ref_strategy:
            agree_str = "[dim]ref[/dim]"
        elif agreements:
            agree_str = f"{sum(agreements) / len(agreements) * 100:.0f}%"
        else:
            agree_str = "n/a"

        success_style = "green" if successes == n_workbooks else "yellow" if successes > 0 else "red"
        table.add_row(
            strat,
            f"{avg_time:.1f}s",
            f"[{success_style}]{successes}/{n_workbooks}[/{success_style}]",
            f"{avg_boxes:.1f}" if box_counts else "-",
            f"{avg_items:.1f}" if item_counts else "-",
            agree_str,
        )

    console.print()
    console.print(table)
    console.print("  Agreement: Jaccard similarity of (item, box_idx, qty) tuples vs sonnet-low.")
    console.print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Benchmark LLM extraction strategies on historical workbooks."
    )
    parser.add_argument(
        "n",
        type=int,
        help="Number of largest workbooks to test (0 = all in target set)",
    )
    parser.add_argument(
        "--strategies",
        type=str,
        default=None,
        help=f"Comma-separated strategies (default: all). Options: {','.join(ALL_STRATEGIES)}",
    )
    parser.add_argument(
        "--offers",
        type=str,
        default=None,
        help="Comma-separated offer IDs to test (overrides N selection)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Ignore cache and re-run all extractions",
    )
    args = parser.parse_args()

    strategies = ALL_STRATEGIES
    if args.strategies:
        strategies = [s.strip() for s in args.strategies.split(",")]
        invalid = [s for s in strategies if s not in STRATEGY_RUNNERS]
        if invalid:
            print(f"Unknown strategies: {invalid}")
            print(f"Available: {', '.join(ALL_STRATEGIES)}")
            sys.exit(1)

    offer_ids = None
    if args.offers:
        try:
            offer_ids = {int(x.strip()) for x in args.offers.split(",")}
        except ValueError:
            parser.error("--offers must be comma-separated integers")

    # Verify claude CLI is available before doing any work
    import shutil

    if not shutil.which("claude"):
        print("Error: 'claude' binary not found on PATH. Install it or add to PATH.")
        sys.exit(1)

    workbooks = select_workbooks(args.n, offer_ids)
    if not workbooks:
        print("No workbooks found.")
        sys.exit(1)

    print(f"\nRunning {len(strategies)} strategies on {len(workbooks)} workbook(s)...")
    print(f"Strategies: {', '.join(strategies)}")
    print()

    all_results: dict[int, dict[str, StrategyResult]] = {}

    for info in workbooks:
        print(f"--- Offer {info.offer_id} ({info.size_kb:.0f} KB, {len(info.sheet_names)} sheets) ---")
        results: dict[str, StrategyResult] = {}

        for strat in strategies:
            # Check cache
            if not args.force:
                cached = _load_cached(info.offer_id, strat)
                if cached is not None:
                    cached.notes = (cached.notes + " (cached)").strip()
                    results[strat] = cached
                    print(f"  {strat}: cached ({cached.time_seconds:.1f}s)")
                    continue

            print(f"  {strat}: running...", end="", flush=True)
            runner = STRATEGY_RUNNERS[strat]
            try:
                result = runner(info.offer_id, info)
            except Exception as e:
                logger.error(f"Strategy {strat} crashed on offer {info.offer_id}: {e}")
                result = StrategyResult(error=f"crashed: {type(e).__name__}: {e}")
            results[strat] = result
            _save_cached(info.offer_id, strat, result)

            status = "OK" if result.success else f"FAIL ({result.error})"
            print(f" {result.time_seconds:.1f}s — {status}")

        all_results[info.offer_id] = results
        print_per_workbook_table(info.offer_id, results, strategies)

    if len(workbooks) > 1:
        print_aggregate_table(all_results, strategies)


if __name__ == "__main__":
    main()

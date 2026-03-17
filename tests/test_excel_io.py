"""Tests for allocator/excel_io.py — format_output, read_overage_from_xlsx."""

import pytest

from allocator.excel_io import format_output, read_overage_from_xlsx
from allocator.models import AllocationResult, CharityBox, MysteryBox


# ── format_output ───────────────────────────────────────────────────────────


class TestFormatOutput:
    def test_merged_box_uses_email(self, make_item, make_box, make_result, make_charity):
        item = make_item(id=1, price=500, pack_order=1)
        box = make_box(name="alice@test.com", merged=True, allocations={1: 2})
        charity = make_charity(name="Charity")
        result = make_result(items=[item], boxes=[box], charity=[charity])
        output = format_output(result)
        lines = output.split("\n")
        assert lines[0] == "ID\talice@test.com\t?Charity"

    def test_standalone_box_has_question_prefix(self, make_item, make_box, make_result):
        item = make_item(id=1, price=500, pack_order=1)
        box = make_box(name="Sm Alex", merged=False, allocations={1: 1})
        result = make_result(items=[item], boxes=[box])
        output = format_output(result)
        lines = output.split("\n")
        assert "?Sm Alex" in lines[0]

    def test_tab_delimited(self, make_item, make_box, make_result):
        item = make_item(id=1, price=500, pack_order=1)
        box = make_box(allocations={1: 3})
        result = make_result(items=[item], boxes=[box])
        output = format_output(result)
        lines = output.split("\n")
        assert len(lines) == 2
        parts = lines[1].split("\t")
        assert parts[0] == "1"
        assert parts[1] == "3"

    def test_empty_qty_is_blank(self, make_item, make_box, make_result):
        items = [
            make_item(id=1, price=500, pack_order=1),
            make_item(id=2, price=300, pack_order=2),
        ]
        box1 = make_box(name="a@test", allocations={1: 1})
        box2 = make_box(name="b@test", allocations={2: 2})
        result = make_result(items=items, boxes=[box1, box2])
        output = format_output(result)
        lines = output.split("\n")
        # Item 1: box1=1, box2="" (no allocation)
        item1_parts = lines[1].split("\t")
        assert item1_parts[0] == "1"
        assert item1_parts[1] == "1"
        assert item1_parts[2] == ""

    def test_sorted_by_pack_order(self, make_item, make_box, make_result):
        items = [
            make_item(id=1, price=500, pack_order=5),
            make_item(id=2, price=300, pack_order=1),
        ]
        box = make_box(allocations={1: 1, 2: 1})
        result = make_result(items=items, boxes=[box])
        output = format_output(result)
        lines = output.split("\n")
        # pack_order=1 (id=2) should come first
        assert lines[1].startswith("2\t")
        assert lines[2].startswith("1\t")

    def test_charity_in_output(self, make_item, make_box, make_result, make_charity):
        item = make_item(id=1, pack_order=1)
        charity = make_charity(name="Test Charity", allocations={1: 3})
        result = make_result(items=[item], charity=[charity])
        output = format_output(result)
        assert "?Test Charity" in output

    def test_no_allocations_header_only(self, make_result, make_box):
        box = make_box()
        result = make_result(boxes=[box])
        output = format_output(result)
        lines = output.split("\n")
        assert len(lines) == 1  # header only


# ── read_overage_from_xlsx ──────────────────────────────────────────────────


class TestReadOverageFromXlsx:
    def test_reads_overage(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Overage"])
        ws.append([101, "Apples", 5])
        ws.append([102, "Bananas", 3])
        ws.append([103, "Broccoli", 0])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        result = read_overage_from_xlsx(path)
        assert result == {101: 5, 102: 3}

    def test_missing_id_column_raises(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Overage"])
        ws.append(["Apples", 5])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="No 'ID' column"):
            read_overage_from_xlsx(path)

    def test_missing_overage_column_raises(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name"])
        ws.append([101, "Apples"])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="No 'Overage' column"):
            read_overage_from_xlsx(path)

    def test_none_overage_treated_as_zero(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Overage"])
        ws.append([101, None])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        result = read_overage_from_xlsx(path)
        assert 101 not in result  # qty=0 is excluded

    def test_non_numeric_id_skipped(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Overage"])
        ws.append(["abc", 5])
        ws.append([101, 3])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        result = read_overage_from_xlsx(path)
        assert result == {101: 3}

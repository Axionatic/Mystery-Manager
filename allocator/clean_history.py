"""
Historical data cleanup script.

Extracts the final allocation tab from each XLSX file, normalizes column
formats, classifies columns, and produces clean CSVs for algorithm validation.

Supports four tiers of data quality:
  A (75-106): Standard 3-tab with IDs and emails — in historical/
  B (55-74):  Standard 3-tab with IDs, all standalone — in historical/ or older/
  C (45-54):  Item names only, no IDs — needs LLM matching
  D (24-44):  Rough formats, per-offer overrides
"""

import csv
import json
import logging
import re
from pathlib import Path

import openpyxl

from allocator.box_parser import classify_box, parse_box_name
from allocator.config import (
    BOX_SIZE_OVERRIDES,
    BUFFER_IDENTIFIERS,
    CHARITY_IDENTIFIERS,
    DONATION_IDENTIFIERS,
    SKIP_COLUMN_IDENTIFIERS,
    STAFF_IDENTIFIERS,
    STANDALONE_NAME_TO_EMAIL,
    STOCK_IDENTIFIERS,
    SUM_IDENTIFIERS,
)

logger = logging.getLogger(__name__)

HISTORICAL_DIR = Path(__file__).parent.parent / "historical"
OLDER_DIR = HISTORICAL_DIR / "older"
CLEANED_DIR = Path(__file__).parent.parent / "cleaned"

# Offers to skip entirely (no usable mystery data)
SKIP_OFFERS = {44}

# Per-offer sheet overrides: {offer_id: sheet_name}
SHEET_OVERRIDES = {}

# Preferred sheet names in priority order
_PREFERRED_SHEETS = [
    "Mystery Trello", "mystery trello",
    "Trello",
    "Mystery Data",
    "Mystery box Trello",
    "Mystery Boxes",
]

# Sheet names to try as fallback (calc tabs — have extra columns but usable)
_FALLBACK_SHEETS = [
    "Mystery Calc", "mystery calc", "Mystery",
    "Trello Mystery",
]


# ---------------------------------------------------------------------------
# Column classification
# ---------------------------------------------------------------------------

def classify_column(header: str | None) -> tuple[str, str | None]:
    """
    Classify a column header into a type.

    Returns:
        (type, size_tier) where type is one of:
        "id", "item", "merged", "standalone", "donation", "charity",
        "staff", "stock", "buffer", "sum", "skip"
        and size_tier is "small", "medium", "large", or None.
    """
    if header is None:
        return ("skip", None)

    header = str(header).strip()
    if not header:
        return ("skip", None)

    if header == "ID":
        return ("id", None)
    if header == "Item":
        return ("item", None)

    # Data columns that sometimes appear in the allocation tab
    if header in ("Price Ea", "Ov - Mys", "Qty Sold", "Required Buy", "Overage"):
        return ("skip", None)

    # Exact match against known special columns
    if header in CHARITY_IDENTIFIERS:
        return ("charity", None)
    if header in STOCK_IDENTIFIERS:
        return ("stock", None)
    if header in BUFFER_IDENTIFIERS:
        return ("buffer", None)
    if header in SUM_IDENTIFIERS:
        return ("sum", None)

    # Check donations first (before standalone check, since some donations
    # use the same Sm/Md prefix pattern)
    if header in DONATION_IDENTIFIERS:
        return ("donation", None)

    # Staff who self-pick (not a standard mystery box)
    if header in STAFF_IDENTIFIERS:
        return ("staff", None)

    # Email pattern → merged mystery box
    if "@" in header:
        # But also check if it's a known donation email
        if header in DONATION_IDENTIFIERS:
            return ("donation", None)
        return ("merged", None)

    # Standalone box pattern: [Sm|Md|Lg] [Name] or [Name] [Sm|Md|Lg]
    size_tier = _infer_size_tier(header)
    return ("standalone", size_tier)


def _infer_size_tier(header: str) -> str | None:
    """Infer box size tier from a standalone column header."""
    if header in BOX_SIZE_OVERRIDES:
        return BOX_SIZE_OVERRIDES[header]

    h = header.lower().strip()

    # Check prefix pattern: "Sm Name", "Md Name", "Lg Name"
    if h.startswith("sm ") or h.startswith("small "):
        return "small"
    if h.startswith("md ") or h.startswith("med ") or h.startswith("medium "):
        return "medium"
    if h.startswith("lg ") or h.startswith("large "):
        return "large"

    # Check suffix pattern: "Name Sm", "Name sm"
    if h.endswith(" sm") or h.endswith(" small"):
        return "small"
    if h.endswith(" md") or h.endswith(" med") or h.endswith(" medium"):
        return "medium"
    if h.endswith(" lg") or h.endswith(" large"):
        return "large"

    # Market Mystery, 4-Sale, etc. - assume small if no size indicator
    if "market" in h or "mystery" in h or "4-sale" in h:
        return "small"

    # Default: unknown size
    return None


def classify_column_extended(header: str | None) -> tuple[str, str | None]:
    """
    Extended column classifier with box_parser for improved box detection.

    Handles older offers' naming conventions (Lge CCI, ?Sm Boatshed, etc.)
    that the basic classify_column misses.
    """
    if header is None:
        return ("skip", None)

    h = str(header).strip()
    if not h:
        return ("skip", None)

    # Structural columns first
    if h == "ID":
        return ("id", None)
    if h in ("Item", "Name"):
        return ("item", None)

    # Extended skip patterns
    if h in SKIP_COLUMN_IDENTIFIERS:
        return ("skip", None)

    # Known special columns
    if h in STOCK_IDENTIFIERS:
        return ("stock", None)
    if h in BUFFER_IDENTIFIERS:
        return ("buffer", None)
    if h in SUM_IDENTIFIERS:
        return ("sum", None)

    # Numeric-only or empty-ish headers are skip columns
    try:
        float(h)
        return ("skip", None)
    except ValueError:
        pass

    # Use box_parser for box-type classification
    cleaned, size_tier, box_type = classify_box(h)

    if box_type == "donation":
        return ("donation", None)
    if box_type == "staff":
        return ("staff", None)
    if box_type == "merged":
        return ("merged", size_tier)

    # Check if this looks like a structural column we missed
    h_lower = h.lower()
    if any(kw in h_lower for kw in ("column", "actual", "total", "difference")):
        return ("skip", None)

    return ("standalone", size_tier)


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def extract_offer_id(filename: str) -> int | None:
    """Extract offer ID from filename like 'offer_104_shopping_list.xlsx'."""
    match = re.search(r"offer_(\d+)", filename)
    return int(match.group(1)) if match else None


def discover_files(*dirs: Path) -> dict[int, tuple[Path, str]]:
    """
    Scan directories for XLSX files and return best file per offer.

    Returns {offer_id: (filepath, source_dir_name)}.
    When duplicates exist across dirs, prefers files in earlier directories.
    """
    candidates: dict[int, list[tuple[Path, str]]] = {}

    for d in dirs:
        if not d.exists():
            continue
        dir_label = d.name if d.name != "historical" else "historical"
        if "older" in str(d):
            dir_label = "historical/older"

        for path in sorted(d.glob("offer_*_shopping_list*.xlsx")):
            offer_id = extract_offer_id(path.name)
            if offer_id is None:
                continue
            candidates.setdefault(offer_id, []).append((path, dir_label))

    result = {}
    for offer_id, paths in candidates.items():
        if len(paths) == 1:
            result[offer_id] = paths[0]
        else:
            # Prefer canonical name, then files in historical/ over older/
            def _score(entry):
                path, source = entry
                canonical = path.name == f"offer_{offer_id}_shopping_list.xlsx"
                in_main = "older" not in source
                has_final = "final" in path.name.lower()
                return (canonical, in_main, has_final, path.stat().st_size)

            best = max(paths, key=_score)
            result[offer_id] = best

    return result


# ---------------------------------------------------------------------------
# Sheet selection
# ---------------------------------------------------------------------------

def _sheet_has_column(ws, col_name: str, max_rows: int = 5) -> int | None:
    """Check if sheet has a column with given name in first max_rows. Returns row number or None."""
    for row_idx in range(1, min(max_rows + 1, (ws.max_row or 0) + 1)):
        for cell in ws.iter_cols(min_row=row_idx, max_row=row_idx):
            val = cell[0].value
            if val is not None and str(val).strip() == col_name:
                return row_idx
    return None


def select_allocation_sheet(wb, offer_id: int):
    """
    Select the best allocation sheet from a workbook.

    Returns (worksheet, sheet_name, header_row) or (None, None, None).
    header_row is 1-indexed.
    """
    sheets = wb.sheetnames

    # Check per-offer override
    if offer_id in SHEET_OVERRIDES:
        override = SHEET_OVERRIDES[offer_id]
        if override in sheets:
            ws = wb[override]
            hr = find_header_row(ws)
            return (ws, override, hr)

    # Try preferred sheet names (Trello tabs — cleanest data)
    for name in _PREFERRED_SHEETS:
        for sn in sheets:
            if sn.lower() == name.lower():
                ws = wb[sn]
                if ws.max_row and ws.max_row >= 2:
                    hr = find_header_row(ws)
                    return (ws, sn, hr)

    # Fall back to last sheet if it has ID or Item column (matches old behavior)
    ws = wb[sheets[-1]]
    if ws.max_row and ws.max_row >= 2:
        hr = find_header_row(ws)
        headers = _read_header_row(ws, hr)
        header_strs = {str(h).strip() for h in headers if h}
        if "ID" in header_strs or "Item" in header_strs or "Name" in header_strs:
            return (ws, sheets[-1], hr)

    # Try fallback sheets (Calc tabs — have extra columns but usable)
    for name in _FALLBACK_SHEETS:
        for sn in sheets:
            if sn.lower() == name.lower():
                ws = wb[sn]
                if ws.max_row and ws.max_row >= 2:
                    hr = find_header_row(ws)
                    return (ws, sn, hr)

    # Try the shopping list sheet if it has box-like columns
    ws = wb[sheets[0]]
    if ws.max_row and ws.max_row >= 2:
        hr = find_header_row(ws)
        headers = _read_header_row(ws, hr)
        # Check for box-like columns (Lge CCI, Med1, M Box 1, etc.)
        box_count = 0
        for h in headers:
            if h is None:
                continue
            _, _, btype = classify_box(str(h).strip())
            if btype in ("standalone", "merged", "donation"):
                box_count += 1
        if box_count >= 3:
            return (ws, sheets[0], hr)

    return (None, None, None)


def find_header_row(ws, max_rows: int = 5) -> int:
    """
    Find the row containing column headers (ID or Item).

    Returns 1-indexed row number. Defaults to 1 if not found.
    """
    for row_idx in range(1, min(max_rows + 1, (ws.max_row or 0) + 1)):
        row_vals = [cell.value for cell in list(ws.iter_rows(
            min_row=row_idx, max_row=row_idx))[0]]
        for val in row_vals:
            if val is not None:
                s = str(val).strip()
                if s in ("ID", "Item", "Name"):
                    return row_idx
    return 1


def _read_header_row(ws, row_num: int = 1) -> list:
    """Read a specific row as header values."""
    return [cell.value for cell in list(ws.iter_rows(
        min_row=row_num, max_row=row_num))[0]]


# ---------------------------------------------------------------------------
# Transposition detection
# ---------------------------------------------------------------------------

def _is_transposed(ws, header_row: int = 1) -> bool:
    """
    Check if a sheet has items as columns and boxes as rows (transposed).

    Heuristic: headers contain many long produce-like names (not emails),
    and column A rows 2+ have short box-like names.
    """
    headers = _read_header_row(ws, header_row)
    if len(headers) < 5:
        return False

    # If any header contains "@", this is a standard sheet with email columns
    for h in headers:
        if h and isinstance(h, str) and "@" in h:
            return False

    # If header has "ID", it's standard orientation
    for h in headers:
        if h and str(h).strip() == "ID":
            return False

    # Check if most header values (cols 1+) look like produce item names
    # (long strings without @ and not matching known column patterns)
    item_like = 0
    for h in headers[1:30]:  # sample first 30 columns
        if h and isinstance(h, str):
            s = h.strip()
            if len(s) > 15 and "@" not in s:
                item_like += 1

    if item_like < 5:
        return False

    # Check if column A values (rows after header) look like box names
    box_like = 0
    numeric_col_a = 0
    for row in ws.iter_rows(min_row=header_row + 1,
                            max_row=min(header_row + 5, ws.max_row or 0),
                            min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val and isinstance(val, str):
            stripped = str(val).strip()
            if len(stripped) < 30:
                box_like += 1
        elif val is not None and isinstance(val, (int, float)):
            numeric_col_a += 1

    if box_like >= 2:
        return True

    # If column A is numeric (row numbers), check column B for box names
    # (e.g., offer 53: "Number", "Box", then item names)
    if numeric_col_a >= 2 and len(headers) > 1:
        box_like_b = 0
        for row in ws.iter_rows(min_row=header_row + 1,
                                max_row=min(header_row + 5, ws.max_row or 0),
                                min_col=2, max_col=2, values_only=True):
            val = row[0]
            if val and isinstance(val, str):
                stripped = str(val).strip()
                if len(stripped) < 30:
                    box_like_b += 1
        if box_like_b >= 2:
            return True

    return False


def _read_transposed(ws, header_row: int = 1):
    """
    Read a transposed sheet (items as columns, boxes as rows).

    Returns (item_names, box_data) where:
        item_names = list of item names (from column headers)
        box_data = {box_header: {item_name: qty}}
    """
    headers = _read_header_row(ws, header_row)

    # Detect "Number"/"Box" pattern (offer 53): box names in col B, items from col C+
    h0 = str(headers[0]).strip().lower() if headers[0] else ""
    h1 = str(headers[1]).strip().lower() if len(headers) > 1 and headers[1] else ""
    if h0 in ("number", "#") and h1 in ("box", "name"):
        box_col = 1
        item_start = 2
    else:
        box_col = 0
        item_start = 1

    # Columns from item_start onward are item names
    item_names = []
    for h in headers[item_start:]:
        if h and str(h).strip():
            item_names.append(str(h).strip())
        else:
            break  # stop at first blank

    # Each subsequent row is a box
    box_data = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        box_name = row[box_col] if box_col < len(row) else None
        if box_name is None:
            continue
        box_name = str(box_name).strip()
        if not box_name:
            continue

        # Skip non-box rows (like "Pack order" metadata)
        if box_name.lower() in ("pack order", "total", "sum"):
            continue

        allocations = {}
        for i, item_name in enumerate(item_names):
            col = item_start + i
            val = row[col] if col < len(row) else None
            if val is not None:
                try:
                    qty = float(val)
                    if qty > 0:
                        allocations[item_name] = int(qty) if qty == int(qty) else qty
                except (ValueError, TypeError):
                    pass

        if allocations:
            box_data[box_name] = allocations

    return item_names, box_data


# ---------------------------------------------------------------------------
# Tier determination
# ---------------------------------------------------------------------------

def _determine_tier(offer_id: int, source_dir: str) -> str:
    """Determine data quality tier based on offer_id and source."""
    if offer_id >= 75:
        return "A"
    if offer_id >= 64:
        return "A"  # Tier A in historical/ dir (with quirks)
    if offer_id >= 55:
        return "B"
    if offer_id >= 45:
        return "C"
    return "D"


# ---------------------------------------------------------------------------
# Processing: with IDs (Tiers A, B)
# ---------------------------------------------------------------------------

def _classify_headers(headers, use_extended: bool = False):
    """
    Classify all column headers.

    Returns (classifications, id_col, item_col, mystery_cols, charity_cols, donation_cols).
    """
    classifier = classify_column_extended if use_extended else classify_column

    classifications = []
    id_col = None
    item_col = None
    mystery_cols = []
    charity_cols = []
    donation_cols = []

    for i, h in enumerate(headers):
        col_type, size_tier = classifier(h)
        classifications.append({
            "index": i,
            "header": str(h) if h else None,
            "type": col_type,
            "size_tier": size_tier,
        })

        if col_type == "id":
            id_col = i
        elif col_type == "item":
            item_col = i
        elif col_type in ("merged", "standalone"):
            mystery_cols.append((i, str(h), col_type, size_tier))
        elif col_type in ("charity", "donation"):
            # Both charity and donation go to charity_cols for CSV output
            if col_type == "charity":
                charity_cols.append((i, str(h)))
            else:
                donation_cols.append((i, str(h)))

    return classifications, id_col, item_col, mystery_cols, charity_cols, donation_cols


def _read_data_rows(ws, header_row, id_col, mystery_cols, charity_cols):
    """Read allocation data from rows with an ID column."""
    rows = []
    charity_rows = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_id = row[id_col]
        if row_id is None:
            continue
        try:
            row_id = int(row_id)
        except (ValueError, TypeError):
            continue

        # Mystery box data
        row_data = {"id": row_id}
        for col_idx, col_header, col_type, _ in mystery_cols:
            val = row[col_idx] if col_idx < len(row) else None
            raw = float(val) if val else 0.0
            row_data[col_header] = int(raw) if raw == int(raw) else raw
        rows.append(row_data)

        # Charity data
        if charity_cols:
            charity_data = {"id": row_id}
            for col_idx, col_header in charity_cols:
                val = row[col_idx] if col_idx < len(row) else None
                raw = float(val) if val else 0.0
                charity_data[col_header] = int(raw) if raw == int(raw) else raw
            charity_rows.append(charity_data)

    return rows, charity_rows


def _process_with_ids(ws, headers, header_row, offer_id, filepath,
                      sheet_name, source_dir, tier):
    """Process a sheet that has an ID column (Tiers A, B)."""
    use_extended = tier in ("B", "C", "D") or offer_id <= 74

    clsf, id_col, item_col, mystery_cols, charity_cols, donation_cols = \
        _classify_headers(headers, use_extended=use_extended)

    if id_col is None:
        return None

    rows, charity_rows = _read_data_rows(
        ws, header_row, id_col, mystery_cols, charity_cols)

    box_names = [h for _, h, _, _ in mystery_cols]
    box_sizes = {h: sz for _, h, _, sz in mystery_cols}

    metadata = {
        "offer_id": offer_id,
        "filename": filepath.name,
        "sheet_name": sheet_name,
        "total_items": len(rows),
        "box_count": len(mystery_cols),
        "box_names": box_names,
        "box_sizes": box_sizes,
        "box_types": {h: t for _, h, t, _ in mystery_cols},
        "charity_names": [h for _, h in charity_cols],
        "donation_names": [h for _, h in donation_cols],
        "classifications": clsf,
        "tier": tier,
        "source_dir": source_dir,
    }

    return {
        "metadata": metadata,
        "mystery_rows": rows,
        "charity_rows": charity_rows,
    }


# ---------------------------------------------------------------------------
# Processing: with names only (Tier C)
# ---------------------------------------------------------------------------

def _process_with_names(ws, headers, header_row, offer_id, filepath,
                        sheet_name, source_dir, tier):
    """Process a sheet with Item/Name column but no IDs (Tier C)."""
    clsf, id_col, item_col, mystery_cols, charity_cols, donation_cols = \
        _classify_headers(headers, use_extended=True)

    if item_col is None:
        logger.warning(f"Offer {offer_id}: no Item column found")
        return None

    if not mystery_cols:
        logger.warning(f"Offer {offer_id}: no box columns found")
        return None

    # Try to match item names to DB IDs
    try:
        from allocator.name_matcher import match_items
    except ImportError:
        logger.warning(f"Offer {offer_id}: name_matcher not available, skipping Tier C")
        return None

    # Collect item names and optional price column
    xlsx_names = []
    price_col = None
    for i, h in enumerate(headers):
        hs = str(h).strip() if h else ""
        if hs in ("JS Price Ea", "Price Ea"):
            price_col = i

    price_data = {}
    name_rows_raw = []  # (item_name, row_tuple)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item_name = row[item_col]
        if item_name is None:
            continue
        item_name = str(item_name).strip()
        if not item_name:
            continue
        xlsx_names.append(item_name)
        name_rows_raw.append((item_name, row))

        if price_col is not None:
            pv = row[price_col]
            if pv is not None:
                try:
                    price_data[item_name] = float(pv)
                except (ValueError, TypeError):
                    pass

    # Run name matching
    mappings = match_items(offer_id, xlsx_names,
                          price_column=price_data if price_data else None)

    # Build rows using matched IDs
    rows = []
    charity_rows = []
    matched = 0
    total = 0

    for item_name, row in name_rows_raw:
        total += 1
        if item_name not in mappings:
            continue
        matched += 1
        item_id = mappings[item_name]["id"]

        row_data = {"id": item_id}
        for col_idx, col_header, col_type, _ in mystery_cols:
            val = row[col_idx] if col_idx < len(row) else None
            raw = float(val) if val else 0.0
            row_data[col_header] = int(raw) if raw == int(raw) else raw
        rows.append(row_data)

        if charity_cols:
            charity_data = {"id": item_id}
            for col_idx, col_header in charity_cols:
                val = row[col_idx] if col_idx < len(row) else None
                raw = float(val) if val else 0.0
                charity_data[col_header] = int(raw) if raw == int(raw) else raw
            charity_rows.append(charity_data)

    match_quality = matched / total if total > 0 else 0.0

    box_names = [h for _, h, _, _ in mystery_cols]
    box_sizes = {h: sz for _, h, _, sz in mystery_cols}

    metadata = {
        "offer_id": offer_id,
        "filename": filepath.name,
        "sheet_name": sheet_name,
        "total_items": len(rows),
        "box_count": len(mystery_cols),
        "box_names": box_names,
        "box_sizes": box_sizes,
        "box_types": {h: t for _, h, t, _ in mystery_cols},
        "charity_names": [h for _, h in charity_cols],
        "donation_names": [h for _, h in donation_cols],
        "classifications": clsf,
        "tier": tier,
        "source_dir": source_dir,
        "name_match_quality": round(match_quality, 3),
    }

    return {
        "metadata": metadata,
        "mystery_rows": rows,
        "charity_rows": charity_rows,
    }


# ---------------------------------------------------------------------------
# Processing: transposed sheets
# ---------------------------------------------------------------------------

def _process_transposed(ws, header_row, offer_id, filepath, sheet_name,
                        source_dir, tier):
    """Process a transposed sheet (items as columns, boxes as rows)."""
    item_names, box_data = _read_transposed(ws, header_row)

    if not box_data:
        logger.warning(f"Offer {offer_id}: transposed sheet has no box data")
        return None

    # Try to match item names to DB IDs
    try:
        from allocator.name_matcher import match_items
        mappings = match_items(offer_id, item_names)
    except ImportError:
        logger.warning(f"Offer {offer_id}: name_matcher not available for transposed")
        return None

    # Build rows: pivot from {box: {item: qty}} to {item_id: {box: qty}}
    item_to_id = {name: m["id"] for name, m in mappings.items()}
    all_item_ids = set()
    for allocations in box_data.values():
        for name in allocations:
            if name in item_to_id:
                all_item_ids.add(item_to_id[name])

    # Classify box names
    mystery_boxes = []
    charity_boxes = []
    donation_boxes = []
    for box_name in box_data:
        cleaned, size_tier, box_type = classify_box(box_name)
        if box_type in ("merged", "standalone"):
            mystery_boxes.append((box_name, box_type, size_tier))
        elif box_type in ("charity", "donation"):
            if box_type == "charity":
                charity_boxes.append(box_name)
            else:
                donation_boxes.append(box_name)

    if not mystery_boxes:
        logger.warning(f"Offer {offer_id}: no mystery boxes in transposed data")
        return None

    # Build row-oriented data (same format as _process_with_ids output)
    rows = []
    charity_rows = []
    for item_id in sorted(all_item_ids):
        # Find the item name for this ID
        item_name = None
        for name, mid in item_to_id.items():
            if mid == item_id:
                item_name = name
                break
        if item_name is None:
            continue

        row_data = {"id": item_id}
        for box_name, _, _ in mystery_boxes:
            qty = box_data.get(box_name, {}).get(item_name, 0)
            row_data[box_name] = int(qty) if qty == int(qty) else qty
        rows.append(row_data)

        if charity_boxes:
            charity_data = {"id": item_id}
            for box_name in charity_boxes:
                qty = box_data.get(box_name, {}).get(item_name, 0)
                charity_data[box_name] = int(qty) if qty == int(qty) else qty
            charity_rows.append(charity_data)

    box_names = [bn for bn, _, _ in mystery_boxes]
    box_sizes = {bn: sz for bn, _, sz in mystery_boxes}

    match_quality = len(item_to_id) / len(item_names) if item_names else 0.0

    metadata = {
        "offer_id": offer_id,
        "filename": filepath.name,
        "sheet_name": sheet_name,
        "total_items": len(rows),
        "box_count": len(mystery_boxes),
        "box_names": box_names,
        "box_sizes": box_sizes,
        "box_types": {bn: t for bn, t, _ in mystery_boxes},
        "charity_names": charity_boxes,
        "donation_names": donation_boxes,
        "classifications": [],
        "tier": tier,
        "source_dir": source_dir,
        "name_match_quality": round(match_quality, 3),
    }

    return {
        "metadata": metadata,
        "mystery_rows": rows,
        "charity_rows": charity_rows,
    }


# ---------------------------------------------------------------------------
# Processing: minimal (Tier D shopping list with embedded boxes)
# ---------------------------------------------------------------------------

def _process_minimal(ws, headers, header_row, offer_id, filepath,
                     sheet_name, source_dir, tier):
    """Best-effort processing for Tier D sheets with box columns in shopping list."""
    # Use extended classifier which handles Lge CCI, Med1, M Box 1, etc.
    clsf, id_col, item_col, mystery_cols, charity_cols, donation_cols = \
        _classify_headers(headers, use_extended=True)

    if not mystery_cols:
        logger.warning(f"Offer {offer_id}: no box columns found in minimal sheet")
        return None

    has_id = id_col is not None
    has_item = item_col is not None

    if has_id:
        return _process_with_ids(ws, headers, header_row, offer_id, filepath,
                                sheet_name, source_dir, tier)

    if has_item:
        return _process_with_names(ws, headers, header_row, offer_id, filepath,
                                  sheet_name, source_dir, tier)

    logger.warning(f"Offer {offer_id}: minimal sheet has no ID or Item column")
    return None


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def process_file(filepath: Path, source_dir: str = "historical") -> dict | None:
    """
    Process a single XLSX file and extract allocation data.

    Returns metadata dict with column classifications and data,
    or None if the file can't be processed.
    """
    offer_id = extract_offer_id(filepath.name)
    if offer_id is None:
        print(f"  Skipping {filepath.name}: can't extract offer ID")
        return None

    if offer_id in SKIP_OFFERS:
        print(f"  Skipping offer {offer_id}: in SKIP_OFFERS")
        return None

    tier = _determine_tier(offer_id, source_dir)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Select the best allocation sheet
    ws, sheet_name, header_row = select_allocation_sheet(wb, offer_id)
    if ws is None:
        print(f"  Skipping offer {offer_id}: no suitable sheet found")
        return None

    if ws.max_row is None or ws.max_row < 2:
        print(f"  Skipping offer {offer_id}: sheet '{sheet_name}' has < 2 rows")
        return None

    # Check for transposed orientation
    if _is_transposed(ws, header_row):
        print(f"  Offer {offer_id}: detected transposed sheet '{sheet_name}'")
        return _process_transposed(ws, header_row, offer_id, filepath,
                                   sheet_name, source_dir, tier)

    # Read headers
    headers = _read_header_row(ws, header_row)

    # Determine processing mode
    has_id = any(str(h).strip() == "ID" for h in headers if h)
    has_item = any(str(h).strip() in ("Item", "Name") for h in headers if h)

    if has_id:
        return _process_with_ids(ws, headers, header_row, offer_id, filepath,
                                sheet_name, source_dir, tier)
    elif has_item:
        return _process_with_names(ws, headers, header_row, offer_id, filepath,
                                  sheet_name, source_dir, tier)
    else:
        return _process_minimal(ws, headers, header_row, offer_id, filepath,
                               sheet_name, source_dir, tier)


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def write_mystery_csv(offer_id: int, rows: list[dict], output_dir: Path) -> Path:
    """Write normalized mystery box allocation CSV."""
    if not rows:
        return None

    filepath = output_dir / f"offer_{offer_id}_mystery.csv"
    fieldnames = list(rows[0].keys())

    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return filepath


def write_charity_csv(offer_id: int, rows: list[dict], output_dir: Path) -> Path:
    """Write normalized charity allocation CSV."""
    if not rows:
        return None

    # Only write if there's actual charity data (not just ID + all zeros)
    has_data = any(
        any(v for k, v in row.items() if k != "id")
        for row in rows
    )
    if not has_data:
        return None

    filepath = output_dir / f"offer_{offer_id}_charity.csv"
    fieldnames = list(rows[0].keys())

    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return filepath


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def clean_all(historical_dir: Path = None, output_dir: Path = None,
              include_older: bool = True) -> dict:
    """
    Process all historical XLSX files and produce clean CSVs.

    Args:
        historical_dir: Primary directory (default: historical/)
        output_dir: Output directory for CSVs (default: cleaned/)
        include_older: Also scan historical/older/ (default: True)

    Returns summary dict with metadata for all offers.
    """
    if historical_dir is None:
        historical_dir = HISTORICAL_DIR
    if output_dir is None:
        output_dir = CLEANED_DIR

    output_dir.mkdir(parents=True, exist_ok=True)

    # Discover files across directories
    dirs = [historical_dir]
    older_dir = historical_dir / "older"
    if include_older and older_dir.exists():
        dirs.append(older_dir)

    files = discover_files(*dirs)
    print(f"Found {len(files)} offers across {len(dirs)} directories")

    summary = {"offers": {}}

    for offer_id in sorted(files.keys()):
        filepath, source_dir = files[offer_id]
        print(f"\nProcessing offer {offer_id} ({source_dir})...")

        result = process_file(filepath, source_dir=source_dir)
        if result is None:
            continue

        meta = result["metadata"]

        # Write CSVs
        mystery_path = write_mystery_csv(offer_id, result["mystery_rows"], output_dir)
        charity_path = write_charity_csv(offer_id, result["charity_rows"], output_dir)

        if mystery_path:
            tier_label = f"[{meta.get('tier', '?')}]"
            print(f"  {tier_label} Mystery CSV: {mystery_path.name} "
                  f"({meta['box_count']} boxes, {meta['total_items']} items)")
            if "name_match_quality" in meta:
                print(f"  Name match quality: {meta['name_match_quality']:.0%}")
        if charity_path:
            print(f"  Charity CSV: {charity_path.name}")

        # Print column classification summary
        types_found = {}
        for c in meta.get("classifications", []):
            t = c["type"]
            types_found.setdefault(t, []).append(c["header"])
        for t, cols in sorted(types_found.items()):
            if t not in ("id", "item", "skip"):
                print(f"  {t}: {cols}")

        summary["offers"][str(offer_id)] = meta

    # Write summary JSON
    summary_path = output_dir / "summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"\nSummary written to {summary_path}")
    print(f"Processed {len(summary['offers'])} offers total")

    return summary


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Clean historical XLSX files")
    parser.add_argument("--no-older", action="store_true",
                       help="Skip historical/older/ directory")
    args = parser.parse_args()

    clean_all(include_older=not args.no_older)

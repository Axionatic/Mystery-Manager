"""
Rich-based TUI for reviewing and editing mystery box configuration.

Allows the admin to:
- Review auto-detected boxes with their preferences and notes
- Toggle merged/unmerged per box
- Add off-system boxes (name, size tier)
- Add/edit item exclusion rules per box
- Remove boxes
- Run strategy leaderboard / fill workbook
- Confirm and proceed to allocation
"""

import copy
import sys
import time
from pathlib import Path

from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.text import Text

from allocator.config import BOX_TIERS
from allocator.models import ExclusionRule, MysteryBox

console = Console()


def review_boxes(
    boxes: list[MysteryBox],
    offer_id: int | None = None,
    xlsx_path: Path | None = None,
    charity_names: list[str] | None = None,
) -> list[MysteryBox]:
    """
    Interactive TUI loop for reviewing and editing mystery box configuration.

    Returns the (potentially modified) list of boxes.
    """
    has_context = offer_id is not None and xlsx_path is not None
    while True:
        _display_boxes(boxes)
        console.print()
        console.print("[bold]Commands:[/]")
        console.print("  [cyan]c[/] - Confirm and proceed to allocation")
        console.print("  [cyan]t N[/] - Toggle merged/unmerged for box N")
        console.print("  [cyan]s N TIER[/] - Change tier for box N (small/medium/large)")
        console.print("  [cyan]e N PATTERN[/] - Add exclusion rule for box N")
        console.print("  [cyan]d N[/] - Delete exclusion rules for box N")
        console.print("  [cyan]r N[/] - Remove box N")
        console.print("  [cyan]a NAME TIER[/] - Add off-system box (e.g. 'a \"Sm Rohani\" small')")
        if has_context:
            console.print("  [cyan]l[/] - Run all strategies and show leaderboard")
            console.print("  [cyan]f[/] - Fill workbook with strategy tabs")
        console.print("  [cyan]q[/] - Quit without allocating")
        console.print()

        try:
            cmd = console.input("[bold green]> [/]").strip()
        except (EOFError, KeyboardInterrupt):
            console.print("\nAborted.")
            sys.exit(0)

        if not cmd:
            continue

        parts = cmd.split(None, 2)
        action = parts[0].lower()

        if action == "c":
            console.print("[green]Proceeding with allocation...[/]")
            return boxes

        elif action == "q":
            console.print("Aborted.")
            sys.exit(0)

        elif action == "t" and len(parts) >= 2:
            idx = _parse_index(parts[1], boxes)
            if idx is not None:
                boxes[idx].merged = not boxes[idx].merged
                state = "merged" if boxes[idx].merged else "standalone"
                console.print(f"Box {idx} is now {state}")

        elif action == "s" and len(parts) >= 3:
            idx = _parse_index(parts[1], boxes)
            tier = parts[2].lower()
            if idx is not None and tier in BOX_TIERS:
                boxes[idx].tier = tier
                boxes[idx].target_value = BOX_TIERS[tier]["target_value"]
                console.print(f"Box {idx} tier set to {tier}")
            elif tier not in BOX_TIERS:
                console.print(f"[red]Unknown tier: {tier}. Use small/medium/large[/]")

        elif action == "e" and len(parts) >= 3:
            idx = _parse_index(parts[1], boxes)
            if idx is not None:
                pattern = parts[2]
                boxes[idx].exclusions.append(
                    ExclusionRule(pattern=pattern, source="manual")
                )
                console.print(f"Added exclusion '{pattern}' to box {idx}")

        elif action == "d" and len(parts) >= 2:
            idx = _parse_index(parts[1], boxes)
            if idx is not None:
                manual = [e for e in boxes[idx].exclusions if e.source != "manual"]
                removed = len(boxes[idx].exclusions) - len(manual)
                boxes[idx].exclusions = manual
                console.print(f"Removed {removed} manual exclusion(s) from box {idx}")

        elif action == "r" and len(parts) >= 2:
            idx = _parse_index(parts[1], boxes)
            if idx is not None:
                removed = boxes.pop(idx)
                console.print(f"Removed box: {removed.name}")

        elif action == "a" and len(parts) >= 2:
            # Parse: a "Name With Spaces" tier  OR  a Name tier
            remaining = cmd[2:].strip()
            name, tier = _parse_add_args(remaining)
            if name and tier in BOX_TIERS:
                tier_config = BOX_TIERS[tier]
                new_box = MysteryBox(
                    name=name,
                    tier=tier,
                    merged=False,
                    target_value=tier_config["target_value"],
                )
                boxes.append(new_box)
                boxes.sort(key=lambda b: b.target_value)
                console.print(f"Added {tier} box: {name}")
            else:
                console.print("[red]Usage: a NAME TIER (e.g. a \"Sm Rohani\" small)[/]")

        elif action == "l":
            if not has_context:
                console.print("[red]Leaderboard not available (missing offer_id/xlsx_path)[/]")
            else:
                _run_leaderboard(boxes, offer_id, xlsx_path, charity_names)

        elif action == "f":
            if not has_context:
                console.print("[red]Fill workbook not available (missing offer_id/xlsx_path)[/]")
            else:
                _fill_workbook(boxes, offer_id, xlsx_path, charity_names)

        else:
            console.print(f"[red]Unknown command: {cmd}[/]")


def _display_boxes(boxes: list[MysteryBox]) -> None:
    """Display current box configuration as a rich table."""
    table = Table(title="Mystery Boxes", show_lines=True)
    table.add_column("#", style="dim", width=3)
    table.add_column("Name", style="cyan")
    table.add_column("Tier", style="magenta")
    table.add_column("Type", style="blue")
    table.add_column("Target", justify="right", style="green")
    table.add_column("Preference", style="yellow")
    table.add_column("Exclusions", style="red")
    table.add_column("Notes", max_width=40)

    for i, box in enumerate(boxes):
        box_type = "merged" if box.merged else "standalone"
        pref = box.preference or "-"
        excl = ", ".join(
            e.pattern for e in box.exclusions if e.source == "manual"
        ) or "-"
        notes = (box.notes or "-")[:40]
        target = f"${box.target_value/100:.2f}"

        table.add_row(
            str(i),
            box.name,
            box.tier,
            box_type,
            target,
            pref,
            excl,
            notes,
        )

    console.print(table)


def _parse_index(s: str, boxes: list) -> int | None:
    """Parse and validate a box index."""
    try:
        idx = int(s)
        if 0 <= idx < len(boxes):
            return idx
        console.print(f"[red]Index {idx} out of range (0-{len(boxes)-1})[/]")
    except ValueError:
        console.print(f"[red]Invalid index: {s}[/]")
    return None


def _parse_add_args(s: str) -> tuple[str | None, str]:
    """Parse 'Name tier' or '"Name With Spaces" tier' from add command."""
    s = s.strip()
    if s.startswith('"'):
        # Quoted name
        end = s.find('"', 1)
        if end == -1:
            return None, ""
        name = s[1:end]
        rest = s[end + 1:].strip()
        tier = rest.lower() if rest else "small"
    else:
        parts = s.rsplit(None, 1)
        if len(parts) == 2 and parts[1].lower() in BOX_TIERS:
            name = parts[0]
            tier = parts[1].lower()
        elif len(parts) == 1:
            name = parts[0]
            tier = "small"
        else:
            name = s
            tier = "small"

    return name, tier


def _run_all_strategies(
    boxes: list[MysteryBox],
    offer_id: int,
    xlsx_path: Path,
    charity_names: list[str] | None,
) -> dict[str, tuple]:
    """
    Run all strategies against the current box config.

    Returns dict[strategy_name -> (AllocationResult, elapsed_seconds)].
    """
    from allocator.allocator import allocate
    from allocator.strategies import list_strategies

    strategies = list_strategies()
    # discard-worst before local-search for bootstrap
    ordered = [s for s in strategies if s != "local-search"]
    ordered.append("local-search")

    results = {}
    dw_allocations = None

    for strat in ordered:
        with console.status(f"Running [bold]{strat}[/]..."):
            t0 = time.monotonic()
            try:
                kwargs = {
                    "strategy": strat,
                    "boxes": copy.deepcopy(boxes),
                    "charity_names": charity_names or ["CCI"],
                }
                if strat == "local-search" and dw_allocations is not None:
                    kwargs["bootstrap_allocations"] = dw_allocations
                result = allocate(offer_id, xlsx_path, **kwargs)
            except Exception as e:
                console.print(f"  [red]{strat}: FAILED — {e}[/]")
                continue
            elapsed = time.monotonic() - t0

        if strat == "discard-worst":
            dw_allocations = [dict(box.allocations) for box in result.boxes]

        results[strat] = (result, elapsed)
        console.print(f"  {strat:<20} ({elapsed:.1f}s)")

    return results


def _run_leaderboard(
    boxes: list[MysteryBox],
    offer_id: int,
    xlsx_path: Path,
    charity_names: list[str] | None,
) -> None:
    """Run all strategies and display a leaderboard table."""
    from compare import (
        build_item_lookup,
        compute_available_tags,
        compute_box_metrics,
        compute_composite_score,
    )

    console.print(f"\n[bold]Running all strategies for offer {offer_id}...[/]")
    strategy_results = _run_all_strategies(boxes, offer_id, xlsx_path, charity_names)

    if not strategy_results:
        console.print("[red]No strategies completed successfully.[/]")
        return

    item_lookup = build_item_lookup(offer_id)
    avail_tags = compute_available_tags(item_lookup)

    # Score each strategy
    scored = {}
    for strat, (result, elapsed) in strategy_results.items():
        metrics = []
        for box in result.boxes:
            m = compute_box_metrics(
                box.name, box.allocations, item_lookup, box.tier,
                preference=box.preference, available_tags=avail_tags,
            )
            if m:
                metrics.append(m)
        comp = compute_composite_score(metrics)
        scored[strat] = (comp, elapsed)

    # Build Rich table
    ranked = sorted(scored.items(), key=lambda x: x[1][0]["score"], reverse=True)

    table = Table(title=f"Offer {offer_id} — Strategy Leaderboard", show_lines=False)
    table.add_column("Rank", style="dim", justify="right", width=4)
    table.add_column("Strategy", style="cyan", width=20)
    table.add_column("Score", justify="right", style="bold green", width=7)
    table.add_column("Value", justify="right", width=8)
    table.add_column("Dupes", justify="right", width=8)
    table.add_column("Diver", justify="right", width=8)
    table.add_column("Fair", justify="right", width=8)
    table.add_column("Pref", justify="right", width=8)
    table.add_column("Time", justify="right", style="dim", width=6)

    for i, (name, (comp, elapsed)) in enumerate(ranked, 1):
        table.add_row(
            str(i),
            name,
            f"{comp['score']:.1f}",
            f"{-comp['value_pen']:+.1f}",
            f"{-comp['dupe_pen']:+.1f}",
            f"{-comp['diversity_pen']:+.1f}",
            f"{-comp['fair_pen']:+.1f}",
            f"{-comp['pref_pen']:+.1f}",
            f"{elapsed:.1f}s",
        )

    console.print()
    console.print(table)
    console.print()


def _fill_workbook(
    boxes: list[MysteryBox],
    offer_id: int,
    xlsx_path: Path,
    charity_names: list[str] | None,
) -> None:
    """Run all strategies and fill the XLSX workbook with strategy tabs."""
    import openpyxl
    from fill_workbook import copy_sheet, fill_strategy_sheet

    console.print(f"\n[bold]Running all strategies for offer {offer_id}...[/]")
    strategy_results = _run_all_strategies(boxes, offer_id, xlsx_path, charity_names)

    if not strategy_results:
        console.print("[red]No strategies completed successfully.[/]")
        return

    console.print(f"\n[bold]Writing strategy tabs to {xlsx_path}...[/]")
    wb = openpyxl.load_workbook(xlsx_path)
    template_name = wb.sheetnames[1]  # Sheet1 = allocation template
    template_ws = wb[template_name]
    xlsx_headers = [cell.value for cell in list(template_ws.iter_rows(min_row=1, max_row=1))[0]]

    for strat, (result, _) in strategy_results.items():
        sheet_name = strat[:31]  # Excel sheet name limit
        # Remove existing sheet with this name if present
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = copy_sheet(wb, template_name, sheet_name)
        fill_strategy_sheet(ws, result, xlsx_headers)

    wb.save(xlsx_path)
    console.print(f"[green]Saved {len(strategy_results)} strategy sheets to {xlsx_path}[/]")

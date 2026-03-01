"""
Excel I/O for mystery box allocation.

Reads the tweaked shopping list XLSX to extract overage data.
Writes tab-delimited output for the admin tool textarea.
"""

import logging
from pathlib import Path

import openpyxl

from allocator.models import AllocationResult, CharityBox, MysteryBox

logger = logging.getLogger(__name__)


def read_overage_from_xlsx(filepath: Path) -> dict[int, int]:
    """
    Read item overage quantities from the shopping list XLSX.

    Finds the 'Overage' and 'ID' columns by header name (robust to column
    reordering). Returns dict of {offer_part_id: overage_qty}.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]  # First sheet is the shopping list

    # Find column indices by header name
    headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]

    id_col = None
    overage_col = None
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str == "ID":
            id_col = i
        elif h_str == "Overage" and overage_col is None:
            # Take the first 'Overage' column (there may be a computed one later)
            overage_col = i

    if id_col is None:
        raise ValueError(f"No 'ID' column found in {filepath.name}")
    if overage_col is None:
        raise ValueError(f"No 'Overage' column found in {filepath.name}")

    logger.info(f"Reading {filepath.name}: ID col={id_col}, Overage col={overage_col}")

    overage = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_id = row[id_col] if id_col < len(row) else None
        raw_overage = row[overage_col] if overage_col < len(row) else None

        if raw_id is None:
            continue

        try:
            item_id = int(raw_id)
        except (ValueError, TypeError):
            continue

        try:
            qty = int(raw_overage) if raw_overage else 0
        except (ValueError, TypeError):
            qty = 0

        if qty > 0:
            overage[item_id] = qty

    logger.info(f"Read {len(overage)} items with overage > 0")
    return overage


def format_output(result: AllocationResult) -> str:
    """
    Format allocation result as tab-delimited text for the admin tool.

    Output format matches parseMysteryBoxInput():
    - First row: ID\\tbox_name_1\\tbox_name_2\\t...
    - Subsequent rows: item_id\\tqty_1\\tqty_2\\t...
    - Merged boxes use email as column header
    - Standalone boxes use name with ? prefix (e.g. "?Sm CCI")
    """
    # Build column headers
    columns = []
    for box in result.boxes:
        if box.merged:
            columns.append(box.name)  # email address
        else:
            columns.append(f"?{box.name}")  # ? prefix for standalone
    for charity in result.charity:
        columns.append(f"?{charity.name}")

    # Header row
    lines = ["ID\t" + "\t".join(columns)]

    # Collect all item IDs that have any allocation
    all_item_ids = set()
    for box in result.boxes:
        all_item_ids.update(box.allocations.keys())
    for charity in result.charity:
        all_item_ids.update(charity.allocations.keys())

    # Sort by pack_order then ID for consistent output
    sorted_ids = sorted(
        all_item_ids,
        key=lambda x: (result.items[x].pack_order if x in result.items else 999, x),
    )

    # Data rows
    for item_id in sorted_ids:
        qtys = []
        for box in result.boxes:
            q = box.allocations.get(item_id, 0)
            qtys.append(str(q) if q else "")
        for charity in result.charity:
            q = charity.allocations.get(item_id, 0)
            qtys.append(str(q) if q else "")
        lines.append(f"{item_id}\t" + "\t".join(qtys))

    return "\n".join(lines)

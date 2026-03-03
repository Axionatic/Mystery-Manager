"""
LLM-based workbook analysis and extraction for non-standard historical offers.

Uses Sonnet to interpret complex/irregular XLSX layouts that the algorithmic
pipeline in clean_history.py can't reliably handle. Extracts per-box allocation
data and returns it in a structured format for downstream processing.

Results are cached in mappings/offer_N_llm_extraction.json.
"""

import json
import logging
from pathlib import Path

import openpyxl

from allocator.claude_cli import call_claude_cli
from allocator.config import CHARITY_NAME

logger = logging.getLogger(__name__)

MAPPINGS_DIR = Path(__file__).parent.parent / "mappings"

# Tier A example: offer 64 "Mystery Trello" sheet (first 10 data rows)
# Used to show the LLM what clean allocation data looks like.
_TIER_A_EXAMPLE = """\
Row\tA\tB\tC\tD\tE\tF\tG\tH\tI\tJ\tK\tL
1\t\tMediums\t\tSmalls\t\t\t\t\t\t\t\tCharity
2\tID\talice@example.com\tMd Alex\tbob@example.com\tcharlie@example.edu.au\tdana@example.com\tevan@example.net.au\tfiona@example.com\towner@example.shop\tSm Pat M\tSm Lee H\tCharity
3\t5050\t\t\t\t\t\t\t\t\t\t\t61
4\t5089\t4\t4\t4\t4\t4\t4\t4\t6\t20\t\t
5\t4966\t4\t4\t2\t2\t2\t2\t2\t\t20\t\t
6\t4986\t2\t2\t2\t2\t2\t2\t2\t2\t\t10\t10
7\t4978\t4\t4\t2\t2\t2\t2\t2\t2\t10\t\t
8\t5090\t\t3\t2\t2\t2\t2\t2\t2\t\t10\t
9\t5004\t\t2\t1\t1\t1\t1\t1\t\t\t\t15
10\t5153\t2\t2\t2\t2\t2\t2\t2\t2\t5\t\t"""


def _format_sheet_as_text(ws, max_rows: int | None = None) -> str:
    """
    Render a worksheet as tab-separated text with row numbers and column letters.

    Args:
        ws: openpyxl worksheet
        max_rows: Limit rows (None = all rows)

    Returns:
        Tab-separated text with "Row\tA\tB\tC..." header.
    """
    if ws.max_row is None or ws.max_row < 1:
        return "(empty sheet)"

    n_rows = min(ws.max_row, max_rows) if max_rows else ws.max_row
    n_cols = ws.max_column or 1

    # Column letter header
    col_letters = []
    for c in range(1, n_cols + 1):
        col_letters.append(openpyxl.utils.get_column_letter(c))

    lines = ["Row\t" + "\t".join(col_letters)]

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=n_rows, values_only=True), 1
    ):
        vals = []
        for v in row:
            if v is None:
                vals.append("")
            elif isinstance(v, float) and v == int(v):
                vals.append(str(int(v)))
            else:
                vals.append(str(v))
        lines.append(f"{r_idx}\t" + "\t".join(vals))

    return "\n".join(lines)


def _build_prompt(offer_id: int, workbook_text: str) -> str:
    """Build the LLM prompt with business context, example, and workbook content."""
    return f"""\
You are analysing a historical spreadsheet from a fruit & veggie box business.

## Business context
Customers order mystery boxes (small, medium, large). We buy bulk boxes from farmers, then split & pack produce into individual customer boxes. Each spreadsheet tracks how items were allocated to boxes for a particular week's offer.

## What good allocation data looks like
Here is an example from offer 64 (a clean "Mystery Trello" sheet):

{_TIER_A_EXAMPLE}

Key features:
- Row 2 has column headers: "ID" column (DB item IDs), then one column per customer box (email or name), then {CHARITY_NAME} (charity)
- Row 1 has optional size grouping labels ("Mediums", "Smalls", "{CHARITY_NAME}")
- Data rows: item ID in first column, then integer quantities per box
- Box naming: emails (merged orders), "Md Name" / "Sm Name" (standalone), "{CHARITY_NAME}" (charity/donation)

## Your task
Analyse ALL sheets in the workbook for offer {offer_id} below. Find the sheet (or region within a sheet) that contains per-box mystery allocation data — i.e., how many of each item goes into each individual customer box.

Important rules:
- Only extract individual customer box columns — NOT totals, subtotals, "Total" columns, price columns, cost columns, or supplier columns
- Each box column should represent ONE customer's box
- Items should be produce items (fruit/vegetables) — skip mystery box line items themselves
- Only include items that have at least 1 non-zero allocation
- Empty/blank cells = 0
- If the data is transposed (items as columns, boxes as rows), handle that
- If there are multiple data tables on one sheet, find the per-box allocation one
- If columns have interspersed "Total" columns between box groups, skip those
- Charity/donation boxes ({CHARITY_NAME}, donation emails) should go in charity_boxes
- If no per-box data exists (only aggregates/summaries), explain in notes

## Workbook content

{workbook_text}

## Required output
Reply with ONLY valid JSON (no markdown fences, no explanation outside the JSON):

{{
  "sheet_used": "sheet name where allocation data was found",
  "data_orientation": "normal" or "transposed",
  "boxes": [
    {{"header": "original column/row header", "name": "cleaned name", "size": "small|medium|large|unknown"}}
  ],
  "items": [
    {{"name": "item name from spreadsheet", "allocations": [qty_per_box_1, qty_per_box_2, ...]}}
  ],
  "charity_boxes": [
    {{"header": "original header", "name": "cleaned name"}}
  ],
  "charity_items": [
    {{"name": "item name", "allocations": [qty_per_charity_box_1, ...]}}
  ],
  "skipped_columns": ["header: reason"],
  "notes": "observations about the data format, any difficulties encountered"
}}

The allocations arrays MUST have exactly the same length as the boxes array (one entry per box, in order). Same for charity_items vs charity_boxes. Use 0 for empty cells."""


def _cache_path(offer_id: int) -> Path:
    return MAPPINGS_DIR / f"offer_{offer_id}_llm_extraction.json"


def _load_cache(offer_id: int) -> dict | None:
    path = _cache_path(offer_id)
    if path.exists():
        with open(path) as f:
            return json.load(f)
    return None


def _save_cache(offer_id: int, data: dict):
    MAPPINGS_DIR.mkdir(parents=True, exist_ok=True)
    with open(_cache_path(offer_id), "w") as f:
        json.dump(data, f, indent=2)


def analyze_and_extract(
    offer_id: int, xlsx_path: Path, force: bool = False
) -> dict | None:
    """
    Use Sonnet to analyze a workbook and extract per-box allocation data.

    Args:
        offer_id: The offer ID
        xlsx_path: Path to the XLSX file
        force: If True, ignore cache and re-extract

    Returns:
        Parsed JSON dict with boxes, items, charity data, and notes.
        None if extraction failed.
    """
    # Check cache
    if not force:
        cached = _load_cache(offer_id)
        if cached is not None:
            logger.info(f"Offer {offer_id}: using cached LLM extraction")
            return cached

    # Read all sheets
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheet_texts = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row is None or ws.max_row < 2:
            sheet_texts.append(f"### Sheet: {sn}\n(empty or 1-row sheet)\n")
            continue
        text = _format_sheet_as_text(ws, max_rows=100)
        sheet_texts.append(f"### Sheet: {sn} ({ws.max_row} rows x {ws.max_column} cols)\n{text}\n")

    workbook_text = "\n".join(sheet_texts)

    # Build prompt and call LLM
    prompt = _build_prompt(offer_id, workbook_text)

    logger.info(f"Offer {offer_id}: calling Sonnet for sheet analysis...")
    raw = call_claude_cli(prompt, model="sonnet", timeout=600, lightweight=True)

    if raw is None:
        logger.warning(f"Offer {offer_id}: LLM call failed")
        return None

    # Parse JSON — strip markdown fences if present
    text = raw.strip()
    if text.startswith("```"):
        # Remove ```json ... ``` wrapper
        lines = text.split("\n")
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines)

    try:
        result = json.loads(text)
    except json.JSONDecodeError as e:
        logger.warning(f"Offer {offer_id}: failed to parse LLM JSON: {e}")
        logger.debug(f"Raw LLM output:\n{raw[:500]}")
        return None

    # Basic validation
    if "boxes" not in result or "items" not in result:
        logger.warning(f"Offer {offer_id}: LLM response missing 'boxes' or 'items'")
        return None

    n_boxes = len(result["boxes"])
    for item in result["items"]:
        if len(item.get("allocations", [])) != n_boxes:
            logger.warning(
                f"Offer {offer_id}: item '{item.get('name')}' has "
                f"{len(item.get('allocations', []))} allocations, expected {n_boxes}"
            )
            # Pad or truncate to match
            allocs = item.get("allocations", [])
            if len(allocs) < n_boxes:
                allocs.extend([0] * (n_boxes - len(allocs)))
            else:
                allocs = allocs[:n_boxes]
            item["allocations"] = allocs

    # Same for charity
    n_charity = len(result.get("charity_boxes", []))
    for item in result.get("charity_items", []):
        allocs = item.get("allocations", [])
        if len(allocs) != n_charity:
            if len(allocs) < n_charity:
                allocs.extend([0] * (n_charity - len(allocs)))
            else:
                allocs = allocs[:n_charity]
            item["allocations"] = allocs

    # Cache result
    _save_cache(offer_id, result)
    logger.info(
        f"Offer {offer_id}: extracted {len(result['items'])} items across "
        f"{n_boxes} boxes from sheet '{result.get('sheet_used', '?')}'"
    )

    return result

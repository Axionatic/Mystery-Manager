#!/usr/bin/env python3
"""
Run all strategies against an offer and write results into the XLSX workbook.

For each strategy, duplicates Sheet1 (the allocation template) and fills in
the allocation quantities. Blank cells for 0.

Usage:
    python3 fill_workbook.py 106 historical/offer_106_shopping_list.xlsx
"""

import sys
import time
from copy import copy
from pathlib import Path


import logging
logging.getLogger("paramiko").setLevel(logging.WARNING)
logging.basicConfig(level=logging.WARNING)

import openpyxl
from openpyxl.utils import get_column_letter

from allocator.allocator import allocate
from allocator.strategies import list_strategies


def copy_sheet(wb, src_name, dst_name):
    """Copy a worksheet, preserving values and formatting."""
    src = wb[src_name]
    dst = wb.copy_worksheet(src)
    dst.title = dst_name
    return dst


def build_box_column_map(result, xlsx_headers):
    """
    Map algorithm boxes/charity to XLSX column indices.

    Returns:
        col_map: {xlsx_col_index: (box_or_charity, is_charity)}
        unmapped_boxes: [(box, is_charity)] - boxes with no matching column
    """
    col_map = {}
    unmapped = []

    # Build lookup of xlsx header -> col index (only customer columns, index 8+)
    header_to_col = {}
    for i, h in enumerate(xlsx_headers):
        if i >= 8 and h:
            header_to_col[h] = i

    # Track which columns are claimed
    claimed_cols = set()

    # Map customer boxes
    # Group algorithm boxes by email to handle duplicates (e.g. rohani has small + large)
    from collections import defaultdict
    boxes_by_name = defaultdict(list)
    for box in result.boxes:
        boxes_by_name[box.name].append(box)

    for email, boxes in boxes_by_name.items():
        if len(boxes) == 1:
            box = boxes[0]
            if email in header_to_col:
                col_map[header_to_col[email]] = (box, False)
                claimed_cols.add(header_to_col[email])
            else:
                unmapped.append((box, False))
        else:
            # Multiple boxes for same email (e.g. rohani: small + large)
            # The email column gets the larger box; look for a "Name Sm" column for smaller
            boxes_sorted = sorted(boxes, key=lambda b: b.target_value, reverse=True)
            large_box = boxes_sorted[0]
            small_boxes = boxes_sorted[1:]

            if email in header_to_col:
                col_map[header_to_col[email]] = (large_box, False)
                claimed_cols.add(header_to_col[email])
            else:
                unmapped.append((large_box, False))

            # Try to match smaller boxes to "FirstName Sm" style columns
            for sbox in small_boxes:
                matched = False
                for h, ci in header_to_col.items():
                    if ci in claimed_cols:
                        continue
                    # Match patterns like "Rohani Sm", "Sandra Sm" etc.
                    h_lower = h.lower()
                    email_prefix = email.split("@")[0].lower()
                    if ("sm" in h_lower or "small" in h_lower) and any(
                        part in h_lower for part in email_prefix[:5].split(".")
                    ):
                        col_map[ci] = (sbox, False)
                        claimed_cols.add(ci)
                        matched = True
                        break
                if not matched:
                    unmapped.append((sbox, False))

    # Map charity boxes
    charity_keywords = {"st andrews", "cci", "charity", "donation", "spc", "spsc"}
    for charity in result.charity:
        matched = False
        for h, ci in header_to_col.items():
            if ci in claimed_cols:
                continue
            if h.lower() in charity_keywords or any(kw in h.lower() for kw in charity_keywords):
                col_map[ci] = (charity, True)
                claimed_cols.add(ci)
                matched = True
                break
        if not matched:
            unmapped.append((charity, True))

    return col_map, unmapped


def fill_strategy_sheet(ws, result, xlsx_headers, item_id_col=6):
    """Fill a worksheet with allocation data from a strategy result."""
    col_map, unmapped = build_box_column_map(result, xlsx_headers)

    # Add columns for unmapped boxes
    next_col = len(xlsx_headers)
    for box, is_charity in unmapped:
        label = f"?{box.name}" if not getattr(box, "merged", False) else box.name
        if is_charity:
            label = f"?{box.name}"
        ws.cell(row=1, column=next_col + 1, value=label)
        col_map[next_col] = (box, is_charity)
        next_col += 1

    # Build item_id -> row mapping from the worksheet
    item_rows = {}
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=item_id_col + 1).value
        if cell_val is not None:
            try:
                item_rows[int(cell_val)] = row_idx
            except (ValueError, TypeError):
                pass

    # Fill in allocations
    for col_idx, (box, is_charity) in col_map.items():
        allocs = box.allocations
        for item_id, qty in allocs.items():
            if item_id in item_rows and qty > 0:
                ws.cell(
                    row=item_rows[item_id],
                    column=col_idx + 1,
                    value=qty,
                )

    # Print mapping summary
    print(f"    Column mapping:")
    for col_idx in sorted(col_map.keys()):
        box, is_charity = col_map[col_idx]
        header = xlsx_headers[col_idx] if col_idx < len(xlsx_headers) else "(new)"
        kind = "charity" if is_charity else getattr(box, "tier", "?")
        n_items = sum(1 for q in box.allocations.values() if q > 0)
        print(f"      Col {get_column_letter(col_idx+1)} {header!r:<35} → {box.name} ({kind}, {n_items} items)")


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 fill_workbook.py <offer_id> <xlsx_path>")
        sys.exit(1)

    offer_id = int(sys.argv[1])
    xlsx_path = Path(sys.argv[2])

    strategies = list_strategies()

    # Run strategies in order (discard-worst before local-search for bootstrap)
    ordered = [s for s in strategies if s != "local-search"]
    ordered.append("local-search")

    # Run all strategies and collect results
    print(f"Running {len(ordered)} strategies against offer {offer_id}...")
    strategy_results = {}
    dw_allocations = None

    for strat in ordered:
        t0 = time.monotonic()
        try:
            kwargs = {"strategy": strat}
            if strat == "local-search" and dw_allocations is not None:
                kwargs["bootstrap_allocations"] = dw_allocations
            result = allocate(offer_id, xlsx_path, **kwargs)
        except Exception as e:
            print(f"  {strat}: FAILED — {e}")
            continue
        elapsed = time.monotonic() - t0
        print(f"  {strat:<20} ({elapsed:.1f}s)")

        if strat == "discard-worst":
            dw_allocations = [dict(box.allocations) for box in result.boxes]

        strategy_results[strat] = result

    # Open workbook and write strategy sheets
    print(f"\nWriting to {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path)
    template_name = wb.sheetnames[1]  # Sheet1 = allocation template

    # Read headers from template
    template_ws = wb[template_name]
    xlsx_headers = [cell.value for cell in list(template_ws.iter_rows(min_row=1, max_row=1))[0]]

    for strat, result in strategy_results.items():
        sheet_name = strat[:31]  # Excel sheet name limit
        print(f"\n  Creating sheet '{sheet_name}':")
        ws = copy_sheet(wb, template_name, sheet_name)
        fill_strategy_sheet(ws, result, xlsx_headers)

    # Save
    wb.save(xlsx_path)
    print(f"\nSaved {len(strategy_results)} strategy sheets to {xlsx_path}")


if __name__ == "__main__":
    main()

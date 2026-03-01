#!/usr/bin/env python3
"""
SUMPRODUCT validation: compare XLSX prices against DB offer_parts.price.

For each offer with a Mystery Calc tab containing Price Ea / JS Price Ea:
  - Compare XLSX price (dollars) vs DB price (cents / 100)
  - Account for pack_size
  - Extract SUMPRODUCT totals per box and compare against DB-derived values
  - Report matches, mismatches, and systematic patterns

Usage:
    python3 validate_prices.py                    # all offers
    python3 validate_prices.py --offers 55,60,74  # specific offers
    python3 validate_prices.py --json             # JSON report output
"""

import argparse
import json
import logging
import sys
from pathlib import Path

import openpyxl

from allocator.clean_history import (
    HISTORICAL_DIR,
    OLDER_DIR,
    discover_files,
    extract_offer_id,
    find_header_row,
)
from allocator.config import detect_pack_size
from allocator.db import fetch_offer_items

logging.getLogger("paramiko").setLevel(logging.WARNING)
logging.basicConfig(level=logging.WARNING)


def _find_calc_sheet(wb):
    """Find the Mystery Calc sheet (has price data)."""
    candidates = ["Mystery Calc", "mystery calc", "Mystery", "Mystery calc"]
    for name in candidates:
        for sn in wb.sheetnames:
            if sn.lower() == name.lower():
                return wb[sn], sn
    # Try the shopping list sheet (some Tier D offers have prices there)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    header_strs = [str(h).strip() if h else "" for h in headers]
    if "JS Price Ea" in header_strs or "Price Ea" in header_strs:
        return ws, wb.sheetnames[0]
    return None, None


def _find_price_columns(headers):
    """Find ID, Item, and price column indices."""
    id_col = item_col = price_col = None
    for i, h in enumerate(headers):
        hs = str(h).strip() if h else ""
        if hs == "ID":
            id_col = i
        elif hs in ("Item", "Name"):
            item_col = i
        elif hs in ("Price Ea", "JS Price Ea"):
            price_col = i
    return id_col, item_col, price_col


def _infer_box_sizes_from_adjacency(box_names, box_sizes):
    """
    Infer unknown box sizes from adjacent known sizes.

    Boxes are typically ordered Large → Medium → Small (left to right).
    """
    sizes = [box_sizes.get(bn) for bn in box_names]
    # Forward pass: if we know a size and the next is unknown, infer
    for i in range(len(sizes) - 1):
        if sizes[i] is not None and sizes[i + 1] is None:
            sizes[i + 1] = sizes[i]
    # Backward pass
    for i in range(len(sizes) - 1, 0, -1):
        if sizes[i] is not None and sizes[i - 1] is None:
            sizes[i - 1] = sizes[i]
    return {bn: sz for bn, sz in zip(box_names, sizes) if sz}


def validate_offer(offer_id: int, filepath: Path) -> dict:
    """
    Validate prices for a single offer.

    Returns report dict with matches, mismatches, and SUMPRODUCT comparison.
    """
    report = {
        "offer_id": offer_id,
        "filepath": str(filepath),
        "has_calc_sheet": False,
        "has_price_column": False,
        "items": [],
        "matches": 0,
        "mismatches": 0,
        "unmatched": 0,
        "sumproduct_totals": {},
    }

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws, sheet_name = _find_calc_sheet(wb)
    if ws is None:
        return report

    report["has_calc_sheet"] = True
    report["sheet_name"] = sheet_name

    header_row = find_header_row(ws)
    headers = [cell.value for cell in list(ws.iter_rows(
        min_row=header_row, max_row=header_row))[0]]

    id_col, item_col, price_col = _find_price_columns(headers)

    if price_col is None:
        return report

    report["has_price_column"] = True

    # Fetch DB items
    db_items = {r["id"]: r for r in fetch_offer_items(offer_id)}

    # Also build name → ID mapping for name-based matching
    name_to_id = {}
    for iid, item in db_items.items():
        name_to_id[item["name"]] = iid

    # Find box columns (for SUMPRODUCT validation)
    from allocator.clean_history import classify_column_extended
    box_cols = []  # (col_idx, box_name, size_tier)
    for i, h in enumerate(headers):
        if h is None or i in (id_col, item_col, price_col):
            continue
        hs = str(h).strip()
        col_type, size_tier = classify_column_extended(hs)
        if col_type in ("merged", "standalone"):
            box_cols.append((i, hs, size_tier))

    # Read data rows
    box_totals = {bn: 0.0 for _, bn, _ in box_cols}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Get item identifier
        item_id = None
        item_name = None

        if id_col is not None:
            raw_id = row[id_col]
            if raw_id is not None:
                try:
                    item_id = int(raw_id)
                except (ValueError, TypeError):
                    pass

        if item_col is not None:
            raw_name = row[item_col]
            if raw_name:
                item_name = str(raw_name).strip()

        # Get XLSX price
        xlsx_price_raw = row[price_col] if price_col < len(row) else None
        if xlsx_price_raw is None:
            continue
        try:
            xlsx_price_dollars = float(xlsx_price_raw)
        except (ValueError, TypeError):
            continue

        xlsx_price_cents = round(xlsx_price_dollars * 100)

        # Resolve to DB item
        db_item = None
        if item_id and item_id in db_items:
            db_item = db_items[item_id]
        elif item_name:
            # Try exact name match
            if item_name in name_to_id:
                item_id = name_to_id[item_name]
                db_item = db_items[item_id]
            else:
                # Try prefix match (XLSX truncates names)
                for db_name, db_id in name_to_id.items():
                    if db_name.startswith(item_name.rstrip(".")):
                        item_id = db_id
                        db_item = db_items[item_id]
                        break

        if db_item is None:
            report["unmatched"] += 1
            continue

        # Compare prices (account for pack size)
        pack_size = detect_pack_size(db_item["name"])
        db_price_cents = db_item["price"]
        db_per_unit = db_price_cents // pack_size

        item_report = {
            "id": item_id,
            "name": db_item["name"],
            "xlsx_price_cents": xlsx_price_cents,
            "db_price_cents": db_price_cents,
            "pack_size": pack_size,
            "db_per_unit_cents": db_per_unit,
        }

        # Check: does XLSX match full price or per-unit price?
        if abs(xlsx_price_cents - db_per_unit) <= 5:
            item_report["match"] = "per_unit"
            report["matches"] += 1
        elif abs(xlsx_price_cents - db_price_cents) <= 5:
            item_report["match"] = "full_pack"
            report["matches"] += 1
        else:
            item_report["match"] = "mismatch"
            item_report["diff_cents"] = xlsx_price_cents - db_per_unit
            report["mismatches"] += 1

        report["items"].append(item_report)

        # Accumulate SUMPRODUCT totals per box
        for col_idx, box_name, _ in box_cols:
            qty_raw = row[col_idx] if col_idx < len(row) else None
            if qty_raw is not None:
                try:
                    qty = float(qty_raw)
                    if qty > 0:
                        box_totals[box_name] += xlsx_price_dollars * qty
                except (ValueError, TypeError):
                    pass

    # Store SUMPRODUCT totals (in cents for consistency)
    report["sumproduct_totals"] = {
        bn: round(total * 100) for bn, total in box_totals.items() if total > 0
    }

    return report


def print_report(reports: list[dict]):
    """Print a summary table of validation results."""
    print(f"\n{'='*70}")
    print(f"{'Offer':>6}  {'Match':>6}  {'Mismatch':>8}  {'Unmatched':>9}  {'Sheet'}")
    print(f"{'='*70}")

    total_match = total_mismatch = total_unmatched = 0

    for r in reports:
        if not r["has_price_column"]:
            print(f"{r['offer_id']:>6}  {'—':>6}  {'—':>8}  {'—':>9}  "
                  f"{'no price column' if r['has_calc_sheet'] else 'no calc sheet'}")
            continue

        print(f"{r['offer_id']:>6}  {r['matches']:>6}  {r['mismatches']:>8}  "
              f"{r['unmatched']:>9}  {r.get('sheet_name', '?')}")

        total_match += r["matches"]
        total_mismatch += r["mismatches"]
        total_unmatched += r["unmatched"]

        # Print mismatches detail
        for item in r["items"]:
            if item["match"] == "mismatch":
                diff = item["diff_cents"]
                print(f"        ↳ {item['name'][:40]:<40} "
                      f"XLSX: ${item['xlsx_price_cents']/100:.2f}  "
                      f"DB: ${item['db_per_unit_cents']/100:.2f}  "
                      f"Δ{diff:+d}¢")

    print(f"\n{'Total':>6}  {total_match:>6}  {total_mismatch:>8}  {total_unmatched:>9}")

    # Print SUMPRODUCT totals for offers that have them
    print(f"\n{'='*70}")
    print("SUMPRODUCT box totals (from XLSX prices × quantities):")
    print(f"{'='*70}")
    for r in reports:
        if r["sumproduct_totals"]:
            print(f"\nOffer {r['offer_id']}:")
            for box_name, total_cents in sorted(r["sumproduct_totals"].items()):
                print(f"  {box_name:<30} ${total_cents/100:.2f}")


def main():
    parser = argparse.ArgumentParser(description="Validate XLSX prices against DB")
    parser.add_argument("--offers", type=str, default=None,
                       help="Comma-separated offer IDs (default: all)")
    parser.add_argument("--json", action="store_true",
                       help="Output JSON report instead of table")
    args = parser.parse_args()

    # Discover files
    files = discover_files(HISTORICAL_DIR, OLDER_DIR)

    if args.offers:
        offer_ids = [int(x.strip()) for x in args.offers.split(",")]
        files = {k: v for k, v in files.items() if k in offer_ids}

    if not files:
        print("No files found to validate")
        sys.exit(1)

    reports = []
    for offer_id in sorted(files.keys()):
        filepath, source_dir = files[offer_id]
        try:
            report = validate_offer(offer_id, filepath)
            reports.append(report)
        except Exception as e:
            print(f"Error processing offer {offer_id}: {e}", file=sys.stderr)

    if args.json:
        # Strip large item lists for readability
        for r in reports:
            r["filepath"] = str(Path(r["filepath"]).name)
        print(json.dumps(reports, indent=2))
    else:
        print_report(reports)


if __name__ == "__main__":
    main()
